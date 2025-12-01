from flask import Blueprint, render_template, request, jsonify, session, current_app, make_response, send_file
from app import db 
from app.models import Conversation, User, Product, Order, OrderItem, ChatSession, CartItem, DealerWiseStockDetails, Customer, FOC, PendingOrderProducts
from app.database_service import DatabaseService
from app.llm_classification_service import LLMClassificationService
from app.web_search_service import WebSearchService
from app.enhanced_order_service import EnhancedOrderService
from app.llm_order_service import LLMOrderService
from app.pricing_service import PricingService
from app.groq_service import GroqService 
from app.email_utils import send_conversation_email
# OTP email removed - authentication not used
# from app.email_utils import send_otp_email
from app.stock_management_service import StockManagementService
from app.azure_search_service import get_search_service
from app.translation_service import get_translation_service
from app.azure_speech_service import get_speech_service
import logging
import time
from datetime import datetime
import json
import os
from io import BytesIO
import re

chatbot_bp = Blueprint('enhanced_chatbot', __name__)

# Single logger initialization - logging.basicConfig should only be called once in __init__.py
logger = logging.getLogger(__name__)

# Initialize services
db_service = None
classification_service = None
web_search_service = None
enhanced_order_service = None
llm_order_service = None
pricing_service = None
llm_service = None

def get_db_service():
    """Get database service instance"""
    global db_service
    if db_service is None:
        db_service = DatabaseService()
    return db_service

def get_classification_service():
    """Get LLM classification service instance"""
    global classification_service
    if classification_service is None:
        classification_service = LLMClassificationService()
    return classification_service

def get_web_search_service():
    """Get web search service instance"""
    global web_search_service
    if web_search_service is None:
        web_search_service = WebSearchService()
    return web_search_service

def get_enhanced_order_service():
    """Get enhanced order service instance"""
    global enhanced_order_service
    if enhanced_order_service is None:
        enhanced_order_service = EnhancedOrderService()
    return enhanced_order_service

def get_llm_order_service():
    """Get LLM order service instance"""
    global llm_order_service
    if llm_order_service is None:
        llm_order_service = LLMOrderService()
    return llm_order_service

def get_pricing_service():
    """Get pricing service instance"""
    global pricing_service
    if pricing_service is None:
        pricing_service = PricingService()
    return pricing_service

def build_product_list_with_foc(products, pricing_service):
    """Helper function to build product list with FOC information"""
    product_list = []
    
    for product in products:
        # Handle both dict (from dealer stock) and Product objects
        if isinstance(product, dict):
            product_code = product.get('product_code', '')
            product_name = product.get('product_name', '')
            sales_price = float(product.get('sales_price', product.get('price', 0)))
            available = int(product.get('available_quantity', 0))
            product_id = product.get('product_id') or product.get('id')
        else:
            # Product object
            price = product.sales_price if hasattr(product, 'sales_price') and product.sales_price else (product.price if hasattr(product, 'price') else 0)
            available = product.available_for_sale if hasattr(product, 'available_for_sale') else 0
            product_code = str(product.id) if not hasattr(product, 'product_code') else product.product_code
            product_name = product.product_name
            sales_price = float(price) if price else 0.0
            available = int(available) if available else 0
            product_id = product.id
        
        # Get FOC information for this product
        # Create a minimal product-like object for FOC lookup
        class ProductLike:
            def __init__(self, pid, pname, pcode):
                self.id = pid
                self.product_name = pname
                self.product_code = pcode
        
        try:
            # Get FOC scheme directly from database
            from app.models import FOC
            
            foc_text = None
            foc_scheme = None
            
            # Try to find FOC by product_id
            if product_id:
                foc_scheme = FOC.query.filter_by(product_id=product_id, is_active=True).first()
            
            # If not found by ID, try by product name matching
            if not foc_scheme and product_name:
                # Try exact match first
                foc_scheme = FOC.query.filter_by(product_name=product_name, is_active=True).first()
                
                # Try normalized matching
                if not foc_scheme:
                    all_foc = FOC.query.filter_by(is_active=True).all()
                    for foc in all_foc:
                        # Simple matching: check if product names contain same key terms
                        product_clean = product_name.upper().split('(')[0].strip()
                        foc_clean = foc.product_name.upper().split('(')[0].strip()
                        if product_clean == foc_clean or product_clean in foc_clean or foc_clean in product_clean:
                            foc_scheme = foc
                            break
            
            # Parse all FOC schemes from database
            foc_schemes_list = []
            if foc_scheme:
                if foc_scheme.scheme_1 and '+' in foc_scheme.scheme_1:
                    parts = foc_scheme.scheme_1.split('+')
                    foc_schemes_list.append({'buy': int(parts[0].strip()), 'free': int(parts[1].strip())})
                
                if foc_scheme.scheme_2 and '+' in foc_scheme.scheme_2:
                    parts = foc_scheme.scheme_2.split('+')
                    foc_schemes_list.append({'buy': int(parts[0].strip()), 'free': int(parts[1].strip())})
                
                if foc_scheme.scheme_3 and '+' in foc_scheme.scheme_3:
                    parts = foc_scheme.scheme_3.split('+')
                    foc_schemes_list.append({'buy': int(parts[0].strip()), 'free': int(parts[1].strip())})
                
                if foc_schemes_list:
                    # For display, show the first scheme (lowest tier)
                    first_scheme = foc_schemes_list[0]
                    foc_text = f"Buy {first_scheme['buy']} Get {first_scheme['free']} Free"
                    logger.info(f"✓ FOC found for {product_name}: {len(foc_schemes_list)} tier(s) - S1:{foc_scheme.scheme_1}, S2:{foc_scheme.scheme_2}, S3:{foc_scheme.scheme_3}")
                else:
                    logger.info(f"No parseable FOC schemes for {product_name}")
            else:
                logger.info(f"No FOC record found for {product_name}")
        except Exception as e:
            logger.warning(f"Could not get FOC for product {product_code}: {str(e)}")
            foc_text = None
            foc_schemes_list = []
        
        product_list.append({
            'product_code': product_code,
            'product_name': product_name,
            'sales_price': sales_price,
            'available_for_sale': available,
            'foc': foc_text,  # Add FOC information for display
            'foc_schemes': foc_schemes_list  # Add all FOC schemes for calculation
        })
    
    return product_list

def get_llm_service():
    """Get LLM service instance (GroqService)"""
    global llm_service
    if llm_service is None:
        llm_service = GroqService()
    return llm_service

def get_stock_management_service():
    """Get stock management service instance"""
    global stock_management_service
    if stock_management_service is None:
        stock_management_service = StockManagementService()
    return stock_management_service

# Initialize stock management service
stock_management_service = None

@chatbot_bp.route('/')
def chat():
    """Enhanced chat interface"""
    # Force reload template from disk (bypass any caching)
    from flask import current_app, request
    import os
    
    # NUCLEAR OPTION: Completely disable Jinja2 caching
    if hasattr(current_app, 'jinja_env'):
        current_app.jinja_env.cache = None
        current_app.jinja_env.auto_reload = True
        # Clear the get_template cache if it exists
        if hasattr(current_app.jinja_env, 'get_template'):
            # Force template reload by updating file timestamp
            template_path = os.path.join(current_app.template_folder, 'enhanced_chat.html')
            if os.path.exists(template_path):
                # Update file modification time to force reload
                os.utime(template_path, None)
                # Also verify file content matches what we expect
                with open(template_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    if 'THREE-MODULE-LOADER' not in content:
                        import logging
                        logging.error(f"CRITICAL: Template file at {template_path} does NOT contain THREE-MODULE-LOADER!")
                        logging.error(f"File contains {len(content)} characters")
    
    # Add cache-busting timestamp to template
    cache_buster = int(time.time() * 1000)  # Milliseconds for better uniqueness
    
    # Log to server console for debugging
    import logging
    logging.info(f"[TEMPLATE-RELOAD] Rendering enhanced_chat.html with cache_buster={cache_buster}")
    logging.info(f"[TEMPLATE-RELOAD] Template folder: {current_app.template_folder}")
    
    # Read template directly to ensure we're using the right file
    template_path = os.path.join(current_app.template_folder, 'enhanced_chat.html')
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            raw_content = f.read()
            logging.info(f"[TEMPLATE-RELOAD] Template file size: {len(raw_content)} bytes")
            logging.info(f"[TEMPLATE-RELOAD] Contains THREE-MODULE-LOADER: {'THREE-MODULE-LOADER' in raw_content}")
            logging.info(f"[TEMPLATE-RELOAD] Contains old code (three.min.js): {'three.min.js' in raw_content or '<script' in raw_content and 'three@0.149' in raw_content and 'querySelector' not in raw_content}")
    
    response = make_response(render_template('enhanced_chat.html', user=None, cache_buster=cache_buster))
    
    # Aggressively disable ALL caching to ensure fresh template
    response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate, max-age=0, private, must-revalidate'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Accel-Expires'] = '0'  # For nginx if used
    response.headers['Last-Modified'] = time.strftime('%a, %d %b %Y %H:%M:%S GMT', time.gmtime())
    
    # Add version header with timestamp - change every time
    version_timestamp = int(time.time())
    response.headers['X-Page-Version'] = f'threejs-v3-{version_timestamp}'
    response.headers['X-Server-Cache-Buster'] = str(cache_buster)
    response.headers['ETag'] = f'"threejs-{version_timestamp}"'  # Quotes for proper ETag format
    
    # Vary header to prevent caching based on different factors
    response.headers['Vary'] = 'Cache-Control'
    
    return response

@chatbot_bp.route('/debug-template')
def debug_template():
    """Debug endpoint to verify template file content"""
    import os
    from flask import current_app
    
    template_path = os.path.join(current_app.template_folder, 'enhanced_chat.html')
    
    result = {
        'template_path': template_path,
        'exists': os.path.exists(template_path),
        'file_size': os.path.getsize(template_path) if os.path.exists(template_path) else 0,
        'contains_three_min_js': False,
        'contains_three_module_loader': False,
        'contains_library_check': False,
        'first_200_chars': ''
    }
    
    if os.path.exists(template_path):
        with open(template_path, 'r', encoding='utf-8') as f:
            content = f.read()
            result['contains_three_min_js'] = 'three.min.js' in content or 'three@0.149' in content
            result['contains_three_module_loader'] = 'THREE-MODULE-LOADER' in content
            result['contains_library_check'] = 'Library Loading Check' in content
            result['first_200_chars'] = content[:200]
    
    return result, 200

@chatbot_bp.route('/static/js/three_avatar.js')
def serve_three_avatar():
    """Serve three_avatar.js with correct MIME type for ES modules"""
    from flask import send_from_directory, current_app, Response
    import os
    
    static_folder = current_app.static_folder
    js_file_path = os.path.join(static_folder, 'js', 'three_avatar.js')
    
    # Read the file and serve with correct MIME type
    if os.path.exists(js_file_path):
        with open(js_file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        response = Response(content, mimetype='application/javascript')
        response.headers['Content-Type'] = 'application/javascript; charset=utf-8'
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        
        return response
    else:
        from flask import abort
        abort(404)

@chatbot_bp.route('/test-vrm')
def test_vrm():
    """Test route to verify VRM file is accessible"""
    from flask import send_from_directory
    import os
    from pathlib import Path
    
    static_path = Path(__file__).resolve().parent.parent / 'static'
    vrm_path = static_path / 'avatars' / 'avatar.vrm'
    
    if vrm_path.exists():
        return send_from_directory(
            str(static_path / 'avatars'),
            'avatar.vrm',
            mimetype='application/octet-stream',
            as_attachment=False
        )
    else:
        return f"VRM file not found at: {vrm_path}", 404


def get_default_action_buttons(user):
    """Get default action buttons based on user type"""
    # Company users get report generation buttons
    if user and user.role == 'company':
        return [
            {'text': 'Generate Report', 'action': 'generate_report'},
            {'text': 'Help', 'action': 'help'}
        ]
    
    # Delivery partners get delivery-specific buttons
    if user and user.role == 'delivery_partner':
        return [
            {'text': 'Track Orders', 'action': 'delivery_dashboard'},
            {'text': 'Help', 'action': 'help'}
        ]
    
    action_buttons = [
        {'text': 'Place Order', 'action': 'place_order'},
        {'text': 'View Open Order', 'action': 'open_order'},
        {'text': 'Company Info', 'action': 'company_info'},
        {'text': 'Product Info', 'action': 'product_info'}
    ]
    
    # For distributors (dealers), check for pending stocks and add button if needed
    if user and user.role == 'distributor':
        try:
            stock_service = get_stock_management_service()
            pending_result = stock_service.get_pending_stock_arrivals(user.unique_id)
            
            if pending_result['success'] and pending_result['count'] > 0:
                # Add "Pending Stocks" button at the beginning
                action_buttons.insert(0, {'text': 'Pending Stocks', 'action': 'pending_stocks'})
        except Exception as e:
            logger.error(f"Error checking pending stocks for action buttons: {str(e)}")
    
    return action_buttons

def translate_response(data: dict, target_language: str = 'en') -> dict:
    """
    Translate response data to target language
    
    Args:
        data: Response dictionary
        target_language: Target language code (en, hi, te, my)
        
    Returns:
        Translated response dictionary
    """
    if target_language == 'en' or not target_language:
        return data
    
    translation_service = get_translation_service()
    if not translation_service.is_available():
        return data
    
    translated_data = data.copy()
    
    # Translate main response text
    if 'response' in translated_data and isinstance(translated_data['response'], str):
        translated_data['response'] = translation_service.translate(
            translated_data['response'], 
            target_language
        )
    
    # Translate action buttons text
    if 'action_buttons' in translated_data and isinstance(translated_data['action_buttons'], list):
        for button in translated_data['action_buttons']:
            if isinstance(button, dict) and 'text' in button:
                button['text'] = translation_service.translate(button['text'], target_language)
    
    # Translate error messages
    if 'error' in translated_data and isinstance(translated_data['error'], str):
        translated_data['error'] = translation_service.translate(
            translated_data['error'], 
            target_language
        )
    
    # Translate product names and descriptions
    if 'products' in translated_data and isinstance(translated_data['products'], list):
        for product in translated_data['products']:
            if isinstance(product, dict):
                if 'name' in product:
                    product['name'] = translation_service.translate(product['name'], target_language)
                if 'description' in product:
                    product['description'] = translation_service.translate(product['description'], target_language)
    
    # Translate order details
    if 'order_details' in translated_data:
        order_details = translated_data['order_details']
        if isinstance(order_details, dict):
            # Translate order status and other text fields
            for key in ['status', 'message', 'error']:
                if key in order_details and isinstance(order_details[key], str):
                    order_details[key] = translation_service.translate(order_details[key], target_language)
    
    return translated_data

def ensure_action_buttons(response_data, user):
    """
    Ensure response has action buttons, adding defaults if missing.
    IMPORTANT: This function translates responses for the user, but the database
    should already have the original English text saved via save_conversation().
    """
    # Get user's language preference from session or request
    target_language = session.get('user_language', 'en')
    
    # Skip adding default buttons for company users if they already have interactive_report_selection
    if isinstance(response_data, tuple):
        # Flask jsonify response (tuple)
        response_obj, status_code = response_data
        if hasattr(response_obj, 'get_json'):
            data = response_obj.get_json()
            if data:
                # Don't add default buttons if interactive components exist
                if data.get('interactive_report_selection'):
                    # Translate before returning (database already has English)
                    data = translate_response(data, target_language)
                    return jsonify(data), status_code
                    
                if 'action_buttons' not in data or not data.get('action_buttons'):
                    data['action_buttons'] = get_default_action_buttons(user)
                
                # Translate response for user (database already has English)
                data = translate_response(data, target_language)
                return jsonify(data), status_code
        return response_data
    elif isinstance(response_data, dict):
        # Dict response
        # Don't add default buttons if interactive components exist
        if response_data.get('interactive_report_selection'):
            return translate_response(response_data, target_language)
            
        if 'action_buttons' not in response_data or not response_data.get('action_buttons'):
            response_data['action_buttons'] = get_default_action_buttons(user)
        
        # Translate response for user (database already has English)
        return translate_response(response_data, target_language)
    
    return response_data

@chatbot_bp.route('/message', methods=['POST'])
def process_message():
    """Process chat message with enhanced HV (Powered by Quantum Blue AI) logic"""
    try:
        from app.input_validation import (
            validate_and_sanitize_message, validate_unique_id,
            MAX_LENGTHS, sanitize_string
        )
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        # Validate and sanitize message
        user_message_raw = data.get('message', '')
        if not isinstance(user_message_raw, str):
            return jsonify({'error': 'Message must be a string'}), 400
        
        user_message = validate_and_sanitize_message(user_message_raw)
        if not user_message:
            error_msg = 'Message cannot be empty or exceeds maximum length'
            user_language = data.get('language', session.get('user_language', 'en'))
            if user_language != 'en':
                translation_service = get_translation_service()
                if translation_service.is_available():
                    error_msg = translation_service.translate(error_msg, user_language)
            return jsonify({'error': error_msg}), 400
        
        # Validate and sanitize language
        user_language_raw = data.get('language', session.get('user_language', 'en'))
        user_language = sanitize_string(str(user_language_raw), max_length=MAX_LENGTHS['language'])
        if not user_language or user_language not in ['en', 'hi', 'my', 'te']:
            user_language = 'en'  # Default to English if invalid
        
        # Store user language preference in session
        session['user_language'] = user_language
        
        # Enhanced onboarding flow with unique ID
        if 'onboarding_state' not in session:
            session['onboarding_state'] = 'ask_unique_id'

        state = session['onboarding_state']

        if 'onboarding' not in session:
            session['onboarding'] = {}

        # Onboarding states
        if state == 'ask_unique_id':
            session['onboarding_state'] = 'get_unique_id'
            response_text = 'Hello! Welcome to HV (Powered by Quantum Blue AI). Please enter your unique ID to continue.'
            # Translate welcome message
            if user_language != 'en':
                translation_service = get_translation_service()
                if translation_service.is_available():
                    response_text = translation_service.translate(response_text, user_language)
            return jsonify({
                'response': response_text
            }), 200

        if state == 'get_unique_id':
            # Handle language change message
            if user_message == 'lang_change':
                return jsonify({'response': 'Language updated'}), 200
            
            # Validate and sanitize unique_id
            unique_id_raw = user_message.strip()
            # validate_unique_id returns boolean, so we check it and use the raw value if valid
            if not validate_unique_id(unique_id_raw):
                response_text = 'Please enter a valid unique ID.'
                if user_language != 'en':
                    translation_service = get_translation_service()
                    if translation_service.is_available():
                        response_text = translation_service.translate(response_text, user_language)
                return jsonify({'response': response_text}), 200
            
            # Use the validated raw value (not the boolean return)
            unique_id = unique_id_raw
            db_service = get_db_service()
            user = db_service.get_user_by_unique_id(unique_id)
            
            if not user:
                response_text = 'Unique ID not found. Please check your ID and try again, or contact support for assistance.'
                if user_language != 'en':
                    translation_service = get_translation_service()
                    if translation_service.is_available():
                        response_text = translation_service.translate(response_text, user_language)
                return jsonify({
                    'response': response_text
                }), 200
            
            if not user.is_active:
                response_text = 'Your account is inactive. Please contact support for assistance.'
                if user_language != 'en':
                    translation_service = get_translation_service()
                    if translation_service.is_available():
                        response_text = translation_service.translate(response_text, user_language)
                return jsonify({
                    'response': response_text
                }), 200
            
            # Set user session
            session['user_id'] = user.id
            session['unique_id'] = user.unique_id
            session['user_type'] = user.role
            session['area'] = user.area  # Use area instead of nearest_warehouse
            session['onboarding_state'] = 'ask_intent'
            
            # Create chat session
            chat_session = db_service.create_chat_session(user.id)
            session['session_id'] = chat_session.session_id
            
            # Generate welcome message with action buttons
            welcome_message = generate_welcome_message(user)
            
            # Get action buttons based on user role
            action_buttons = get_default_action_buttons(user)
            
            # Customize message based on role
            if user.role == 'company':
                full_message = welcome_message  # Company welcome already has instructions
            else:
                full_message = f"{welcome_message}\n\nPlease select an option below:"
            
            # Translate welcome message if needed
            if user_language != 'en':
                translation_service = get_translation_service()
                if translation_service.is_available():
                    full_message = translation_service.translate(full_message, user_language)
            
            return ensure_action_buttons(jsonify({
                'response': full_message,
                'action_buttons': action_buttons,
                'user_info': {
                    'name': user.name,
                    'user_type': user.role,
                    'role': user.role,
                    'warehouse': user.area  # Use area instead of nearest_warehouse
                }
            }), user), 200

        # Ask for user intent after verification
        if state == 'ask_intent':
            # Use LLM to understand what user wants to do
            db_service = get_db_service()
            user = User.query.get(session.get('user_id'))
            
            llm_service = get_llm_service()
            if llm_service and llm_service.client:
                intent_prompt = f"""You are an AI assistant for HV (Powered by Quantum Blue AI). A user just logged in and you asked them what they would like to do.

User's message: "{user_message}"

Based on the user's response, classify their intent into one of these categories:
1. PLACE_ORDER - User wants to place an order (e.g., "place order", "order products", "buy", "purchase", "I want to order")
2. TRACK_ORDER - User wants to track an order (e.g., "track order", "check order status", "where is my order")
3. PRODUCT_INFO - User wants product information (e.g., "show products", "what products do you have", "product list")
4. COMPANY_INFO - User wants company information (e.g., "about company", "company info", "tell me about quantum blue")
5. OTHER - Any other request (general conversation, questions, etc.)

Respond with ONLY a JSON object:
{{
    "intent": "PLACE_ORDER" | "TRACK_ORDER" | "PRODUCT_INFO" | "COMPANY_INFO" | "OTHER",
    "confidence": 0.0-1.0,
    "next_state": "ready" if intent is PLACE_ORDER/TRACK_ORDER/PRODUCT_INFO/COMPANY_INFO, else "continue_conversation"
}}"""

                try:
                    response = llm_service.client.chat.completions.create(
                        model=current_app.config.get('GROQ_MODEL', 'llama-3.3-70b-versatile'),
                        messages=[{"role": "user", "content": intent_prompt}],
                        temperature=0.1,
                        max_tokens=200
                    )
                    
                    result_text = response.choices[0].message.content.strip()
                    # Clean JSON response
                    if result_text.startswith('```'):
                        lines = result_text.split('\n')
                        if len(lines) > 2:
                            result_text = '\n'.join(lines[1:-1])
                    
                    intent_result = json.loads(result_text)
                    detected_intent = intent_result.get('intent', 'OTHER')
                    next_state = intent_result.get('next_state', 'continue_conversation')
                    
                    # Set state based on intent and prepare appropriate response
                    session['onboarding_state'] = 'completed'
                    session['user_intent'] = detected_intent
                    
                    # Initialize context_data for all intents
                    context_data = {}
                    
                    # Generate appropriate response based on intent
                    if detected_intent == 'PLACE_ORDER':
                        # For MRs, check if customer is already selected
                        if user.role == 'mr':
                            # Check if customer is already selected in session
                            if 'selected_customer_id' not in session:
                                # First, ask whether to select existing customer or add new customer
                                response = "To place an order, I need to know which customer you're ordering for. Would you like to select an existing customer or add a new customer?"
                                
                                save_conversation(user.id, user_message, response)
                                
                                return jsonify({
                                    'response': response,
                                    'action_buttons': [
                                        {'text': 'Select Customer', 'action': 'select_customer'},
                                        {'text': 'Add New Customer', 'action': 'add_new_customer'}
                                    ]
                                }), 200
                        
                        # Get actual products from database for dynamic examples
                        products = db_service.get_products_from_dealer_stock(user.area) if user.area else []
                        
                        if products:
                            # Create dynamic example from first 3 products
                            example_products = products[:3]
                            example_text = 'For example: "Order '
                            example_parts = []
                            for i, p in enumerate(example_products):
                                qty = 10 if i == 0 else (5 if i == 1 else 3)  # Vary quantities
                                example_parts.append(f"{qty} {p.product_name}")
                                if i < len(example_products) - 1:
                                    example_parts.append(', ')
                            example_text += ''.join(example_parts) + '"'
                        else:
                            example_text = 'For example: "Order 10 units of product 001, 5 units of 002"'
                        
                        return jsonify({
                            'response': f'Great! I can help you place an order. Please tell me which products you would like to order and their quantities.\n\n{example_text}'
                        }), 200
                    elif detected_intent == 'TRACK_ORDER':
                        # Call handle_track_order directly to show orders table and selection box
                        return handle_track_order(user_message, user, context_data)
                    elif detected_intent == 'PRODUCT_INFO':
                        # Redirect to handle_product_info_or_query for consistent table display
                        return handle_product_info_or_query(user_message, user, context_data)
                    elif detected_intent == 'COMPANY_INFO':
                        return handle_company_info(user_message, user)
                    
                    # For OTHER intent, continue to normal flow
                except Exception as e:
                    logger.error(f"Error classifying intent: {str(e)}", exc_info=True)
                    # Fall through to normal flow
                    session['onboarding_state'] = 'completed'
                    # Initialize context_data for fallback
                    context_data = {}
            else:
                # Fallback: simple keyword matching
                message_lower = user_message.lower()
                session['onboarding_state'] = 'completed'
                if any(kw in message_lower for kw in ['order', 'buy', 'purchase', 'place order', 'i want to place an order']):
                    session['user_intent'] = 'PLACE_ORDER'
                    # For MRs, check if customer is already selected
                    if user.role == 'mr':
                        # Check if customer is already selected in session
                        if 'selected_customer_id' not in session:
                            # First, ask whether to select existing customer or add new customer
                            response = "To place an order, I need to know which customer you're ordering for. Would you like to select an existing customer or add a new customer?"
                            
                            save_conversation(user.id, user_message, response)
                            
                            return jsonify({
                                'response': response,
                                'action_buttons': [
                                    {'text': 'Select Customer', 'action': 'select_customer'},
                                    {'text': 'Add New Customer', 'action': 'add_new_customer'}
                                ]
                            }), 200
                    
                    return jsonify({
                        'response': 'Great! I can help you place an order. Please tell me which products you would like to order and their quantities.\n\nFor example: "Order 50 units of product RB001, 30 units of product RB002, and 100 units of product RB003"'
                    }), 200
                elif any(kw in message_lower for kw in ['track', 'status', 'where is', 'track order', 'i want to track an order']):
                    session['user_intent'] = 'TRACK_ORDER'
                    # Initialize context_data if not already created
                    context_data = {}
                    # Call handle_track_order directly to show orders table and selection box
                    return handle_track_order(user_message, user, context_data)
                else:
                    session['user_intent'] = 'OTHER'

        # Main chat flow (verified user)
        session_user_id = session.get('user_id')
        if not session_user_id:
            # During onboarding, no action buttons needed
            return jsonify({'response': 'Please complete the onboarding process first.', 'action_buttons': []}), 200

        # Get services
        db_service = get_db_service()
        classification_service = get_classification_service()
        web_search_service = get_web_search_service()
        enhanced_order_service = get_enhanced_order_service()
        llm_order_service = get_llm_order_service()
        pricing_service = get_pricing_service()

        # Get user context
        user = User.query.get(session_user_id)
        
        # COMPANY USER FLOW - Handle report generation requests
        if user and user.role == 'company':
            return handle_company_requests(user_message, user)
        
        # Get user's warehouse based on area
        area = user.area if user else None
        warehouse = db_service.get_warehouse_by_area(area) if area else None
        context_data = {
            'user_warehouse': area,
            'user_email': user.email if user else None,
            'user_type': user.role if user else None,
            'user_role': user.role if user else None
        }

        # Get recent orders for context
        recent_orders = db_service.get_orders_by_email(user.email)
        context_data['recent_orders'] = recent_orders[:3]

        # Get conversation history
        conversation_history = db_service.get_conversation_history(session_user_id, limit=10)

        # Check for place order messages FIRST - before any other processing
        # This ensures MRs see customer selection buttons
        message_lower = user_message.lower().strip()
        place_order_keywords = ['i want to place an order', 'place an order', 'place order', 'i want to order', 'order products']
        if any(keyword in message_lower for keyword in place_order_keywords):
            logger.info(f"Detected place order message: {user_message}. User type: {user.role}")
            if user.role == 'mr':
                logger.info(f"MR user detected. Checking for selected_customer_id in session: {'selected_customer_id' in session}")
            return handle_place_order(user_message, user, context_data, conversation_history)

        # Check for simple greetings FIRST (CRITICAL - prevent "hi" from confirming orders)
        # This must come before cart check to avoid misinterpretation
        greeting_keywords = ['hi', 'hello', 'hey', 'good morning', 'good afternoon', 'good evening', 'hola', 'namaste']
        if message_lower in greeting_keywords or any(message_lower.startswith(g) for g in greeting_keywords):
            logger.info(f"Detected greeting message: {user_message}")
            # Check if user has items in cart
            cart_items = db_service.get_cart_items(session_user_id)
            if cart_items:
                cart_count = len(cart_items)
                greeting_response = f"Hello! 👋 You currently have {cart_count} item(s) in your cart. Would you like to:\n• Add more products\n• View your cart\n• Confirm and place your order"
                save_conversation(user.id, user_message, greeting_response)
                return jsonify({
                    'response': greeting_response,
                    'action_buttons': [
                        {'text': 'Add More Products', 'action': 'place_order'},
                        {'text': 'View Cart', 'action': 'view_cart'},
                        {'text': 'Confirm Order', 'action': 'confirm_order'}
                    ]
                }), 200
            else:
                # No cart items - show standard welcome options
                return ensure_action_buttons(jsonify({
                    'response': f"Hello {user.name}! 👋 How can I help you today?"
                }), user), 200
        
        # Check for "select customer", "change customer", and "add new customer" messages BEFORE LLM classification
        # This prevents misclassification when user has items in cart
        # Use more flexible matching for voice input variations
        select_customer_patterns = [
            'select customer', 'select existing customer', 'select a customer',
            'select the customer', 'choose customer', 'pick customer',
            'show customers', 'show customer list', 'customer selection',
            'select an existing customer', 'i want to select customer',
            'i want to select a customer', 'select customer please'
        ]
        change_customer_patterns = [
            'change customer', 'change the customer', 'change my customer',
            'switch customer', 'change to another customer', 'different customer'
        ]
        add_customer_patterns = [
            'add new customer', 'add customer', 'create customer',
            'new customer', 'add a new customer', 'create new customer'
        ]
        
        # Check if message matches any pattern (using 'in' for partial matching)
        message_lower_clean = ' '.join(message_lower.split())  # Normalize whitespace
        
        if any(pattern in message_lower_clean for pattern in select_customer_patterns):
            logger.info(f"Detected select customer message: {user_message}")
            return handle_select_customer(user, context_data)
        elif any(pattern in message_lower_clean for pattern in change_customer_patterns):
            logger.info(f"Detected change customer message: {user_message}")
            # Clear the current selected customer from session
            session.pop('selected_customer_id', None)
            session.pop('selected_customer_unique_id', None)
            return handle_select_customer(user, context_data)
        elif any(pattern in message_lower_clean for pattern in add_customer_patterns):
            logger.info(f"Detected add new customer message: {user_message}")
            return handle_add_new_customer(user, context_data)

        # Use LLM to analyze if user wants to confirm order or add products
        # Get cart items to check if user has items to confirm
        cart_items = db_service.get_cart_items(session_user_id)
        
        # If user has items in cart, use LLM to determine intent
        if cart_items:
            llm_service = get_llm_service()
            if llm_service and llm_service.client:
                try:
                    intent_prompt = f"""You are an AI assistant analyzing user intent in an e-commerce chatbot.

User's message: "{user_message}"

Context:
- The user has items in their shopping cart
- The system just showed them their cart with total price

Determine the user's intent:

1. **CONFIRM_ORDER** - User wants to place/confirm/finalize the order for items already in cart
   Examples: "confirm my order", "place the order", "proceed with order", "yes proceed", "finalize order"
   
2. **ADD_TO_CART** - User wants to add more products to the cart
   Examples: "add 5 quantum sensors", "ok add more items", "put 10 processors", "include product 001"
   
3. **MODIFY_CART** - User wants to modify existing cart items (remove, update quantity)
   Examples: "remove product x", "change quantity", "update cart"

4. **PRODUCT_INFO** - User wants to see product information or list products
   Examples: "list all products", "show products available", "what products do you have", "list products in database"

5. **DATABASE_QUERY** - User wants to query database for specific information
   Examples: "show me all products in database", "list available products", "what's in stock"

6. **GREETING** - User is just greeting (hi, hello, hey, etc.) - NOT confirming order
   Examples: "hi", "hello", "hey", "good morning"

CRITICAL RULES:
- **GREETINGS are NEVER order confirmations**: "hi", "hello", "hey" → GREETING (not CONFIRM_ORDER)
- Simple words like "ok", "yes" ALONE (without "order"/"proceed") → GREETING (NOT CONFIRM_ORDER)
- For CONFIRM_ORDER, user MUST explicitly mention: "order", "confirm", "proceed with order", "place order"
- If message asks to "list", "show", "display" products/database → PRODUCT_INFO or DATABASE_QUERY
- If message contains product names/codes AND action words like "add", "put", "include" → ADD_TO_CART
- NEVER treat product listing requests or greetings as order confirmations

Respond with ONLY a JSON object:
{{
    "intent": "CONFIRM_ORDER" | "ADD_TO_CART" | "MODIFY_CART" | "PRODUCT_INFO" | "DATABASE_QUERY" | "GREETING",
    "confidence": 0.0-1.0,
    "reasoning": "brief explanation"
}}"""

                    response = llm_service.client.chat.completions.create(
                        model=current_app.config.get('GROQ_MODEL', 'llama-3.3-70b-versatile'),
                        messages=[{"role": "user", "content": intent_prompt}],
                        temperature=0.1,
                        max_tokens=200
                    )
                    
                    result_text = response.choices[0].message.content.strip()
                    # Clean JSON response
                    if result_text.startswith('```'):
                        lines = result_text.split('\n')
                        if len(lines) > 2:
                            result_text = '\n'.join(lines[1:-1])
                    
                    intent_result = json.loads(result_text)
                    detected_intent = intent_result.get('intent', 'ADD_TO_CART')
                    confidence = intent_result.get('confidence', 0.5)
                    reasoning = intent_result.get('reasoning', '')
                    
                    logger.info(f"LLM Intent Analysis: {detected_intent} (confidence: {confidence:.2f}) - {reasoning}")
                    
                    # Handle different intents
                    if detected_intent == 'GREETING':
                        # User is just greeting - respond politely and show cart status
                        cart_count = len(cart_items)
                        greeting_response = f"Hello! 👋 You currently have {cart_count} item(s) in your cart. Would you like to:\n• Add more products\n• View your cart\n• Confirm and place your order"
                        save_conversation(user.id, user_message, greeting_response)
                        return jsonify({
                            'response': greeting_response,
                            'action_buttons': [
                                {'text': 'Add More Products', 'action': 'place_order'},
                                {'text': 'View Cart', 'action': 'view_cart'},
                                {'text': 'Confirm Order', 'action': 'confirm_order'}
                            ]
                        }), 200
                    elif detected_intent == 'CONFIRM_ORDER' and confidence >= 0.7:
                        return handle_order_confirmation(user, session_user_id)
                    elif detected_intent in ['PRODUCT_INFO', 'DATABASE_QUERY']:
                        # Handle product listing/database queries
                        logger.info(f"LLM detected {detected_intent} intent - handling database query")
                        return handle_product_info_or_query(user_message, user, context_data)
                    elif detected_intent == 'ADD_TO_CART':
                        # Continue to normal flow to handle adding products
                        logger.info("LLM detected ADD_TO_CART intent - proceeding with order processing")
                except Exception as e:
                    logger.error(f"Error analyzing intent with LLM: {str(e)}")
                    # Fallback to keyword matching if LLM fails
                    message_lower = user_message.lower().strip()
                    
                    # Check for greetings first (CRITICAL - prevent "hi" from confirming orders)
                    greeting_keywords = ['hi', 'hello', 'hey', 'good morning', 'good afternoon', 'good evening', 'hola', 'namaste']
                    if message_lower in greeting_keywords or any(message_lower.startswith(g) for g in greeting_keywords):
                        cart_count = len(cart_items)
                        greeting_response = f"Hello! 👋 You currently have {cart_count} item(s) in your cart. Would you like to:\n• Add more products\n• View your cart\n• Confirm and place your order"
                        save_conversation(user.id, user_message, greeting_response)
                        return jsonify({
                            'response': greeting_response,
                            'action_buttons': [
                                {'text': 'Add More Products', 'action': 'place_order'},
                                {'text': 'View Cart', 'action': 'view_cart'},
                                {'text': 'Confirm Order', 'action': 'confirm_order'}
                            ]
                        }), 200
                    
                    add_keywords = ['add', 'put', 'include']
                    has_add_word = any(keyword in message_lower for keyword in add_keywords)
                    # Check for product codes dynamically instead of hardcoded names
                    import re
                    has_product_mention = bool(re.search(r'\b(rb|bd|qb|product)\s*\d+|\d+\s*(rb|bd|qb|product)\b', message_lower, re.IGNORECASE))
                    
                    # Only confirm if no "add" keyword with product mention AND explicitly mentions order/confirm/proceed
                    if not (has_add_word and has_product_mention):
                        confirmation_keywords = ['confirm order', 'place order', 'proceed with order', 'finalize order', 'yes proceed']
                        if any(keyword in message_lower for keyword in confirmation_keywords):
                            return handle_order_confirmation(user, session_user_id)
        
        # Check for stock confirmation requests for distributors FIRST (before intent classification)
        # This prevents "show pending stocks" from being misclassified as PRODUCT_INFO
        # Also handles "pending orders" as a synonym for "pending stocks"
        if user.role == 'distributor':
            message_lower = user_message.lower().strip()
            stock_keywords = [
                'pending stock', 'pending stocks', 'pending order', 'pending orders',
                'stock arrival', 'confirm stock', 'show stock', 'show pending stock', 
                'show pending stocks', 'show pending orders', 'stock to confirm', 
                'pending arrivals', 'list pending', 'view pending', 'pending items',
                'list pending stock', 'view pending stock', 'display pending stock'
            ]
            if any(keyword in message_lower for keyword in stock_keywords):
                logger.info(f"Detected stock confirmation request: {user_message}")
                return handle_stock_confirmation(user_message, user, context_data)
        
        # Check for "confirm cart" message - show Edit Cart / Place Order options
        message_lower = user_message.lower().strip()
        if message_lower in ['confirm cart', 'confirm my cart', 'cart confirmed']:
            cart_items = db_service.get_cart_items(session_user_id)
            if not cart_items:
                response = "Your cart is empty. Please add some products to your cart first."
                save_conversation(user.id, user_message, response)
                return jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'Place Order', 'action': 'place_order'}
                    ]
                }), 200
            
            # Calculate cart total
            total = sum(item.total_price for item in cart_items)
            response = f"Your cart has been confirmed! You have {len(cart_items)} item(s) in your cart with a total of ${total:,.2f}.\n\nWhat would you like to do next?"
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Edit Cart', 'action': 'edit_cart'},
                    {'text': 'Place Order', 'action': 'place_order_final'}
                ]
            }), 200
        
        # Check for "edit cart" message - show cart editing interface
        if message_lower in ['edit cart', 'edit my cart', 'modify cart', 'change cart']:
            cart_items = db_service.get_cart_items(session_user_id)
            if not cart_items:
                response = "Your cart is empty. There's nothing to edit."
                save_conversation(user.id, user_message, response)
                return jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'Place Order', 'action': 'place_order'}
                    ]
                }), 200
            
            # Return cart items for editing
            cart_data = []
            for item in cart_items:
                product_name = item.product_code
                if hasattr(item, 'product') and item.product:
                    product_name = item.product.product_name
                elif hasattr(item, 'product_name'):
                    product_name = item.product_name
                
                cart_data.append({
                    'id': item.id,
                    'product_code': item.product_code,
                    'product_name': product_name,
                    'quantity': item.quantity,
                    'unit_price': item.unit_price,
                    'total_price': item.total_price
                })
            
            total = sum(item.total_price for item in cart_items)
            response = f"Here are the items in your cart. You can view and edit them using the cart modal.\n\nTotal: ${total:,.2f}"
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'cart_items': cart_data,
                'action_buttons': [
                    {'text': 'View Cart', 'action': 'view_cart'},
                    {'text': 'Place Order', 'action': 'place_order_final'}
                ]
            }), 200
        
        # Check for "place order" message (final confirmation) - actually place the order
        if message_lower in ['place order', 'place my order', 'finalize order', 'confirm order']:
            return handle_order_confirmation(user, session_user_id)
        
        # Check for "Product Info" messages BEFORE intent classification
        # This handles messages like "Product Info" from the action button
        product_info_keywords = ['product info', 'productinfo', 'i want product info', 'show product info']
        if message_lower in product_info_keywords or (message_lower == 'product info'):
            logger.info(f"Detected Product Info button click: {user_message}")
            # Use the dedicated handler function for consistency
            return handle_product_info_or_query(user_message, user, context_data)
        
        # Check for "track order" messages BEFORE intent classification
        # This handles messages like "track order", "I want to track an order" from the action button
        track_order_keywords = ['track order', 'i want to track an order', 'track my order', 'show my orders', 'view my orders']
        if message_lower in track_order_keywords or ('track' in message_lower and 'order' in message_lower):
            logger.info(f"Detected track order message: {user_message}")
            # context_data is already initialized above
            return handle_track_order(user_message, user, context_data)
        
        # Check for "add to cart" messages BEFORE intent classification
        # This handles messages like "add 3 BD-20008 to cart" or "add 3 BD-20008 (Product Name) to cart" from the product selection form
        import re
        add_to_cart_patterns = [
            r'add\s+(\d+)\s+([a-z0-9_-]+)\s+\(([^)]+)\)\s+to\s+cart',  # With product name: "add 3 BD-20008 (Product Name) to cart"
            r'add\s+(\d+)\s+([a-z0-9_-]+)\s+to\s+cart',  # Without product name: "add 3 BD-20008 to cart"
            r'add\s+(\d+)\s+([a-z0-9_-]+)',  # Simple: "add 3 BD-20008"
            r'put\s+(\d+)\s+([a-z0-9_-]+)\s+in\s+cart',
            r'include\s+(\d+)\s+([a-z0-9_-]+)'
        ]
        
        product_name_from_message = None
        for pattern in add_to_cart_patterns:
            match = re.search(pattern, message_lower, re.IGNORECASE)
            if match:
                # This is an "add to cart" message - extract quantity and product code directly
                quantity = int(match.group(1))
                product_code = match.group(2).upper().strip()  # Normalize product code
                # Check if product name was included in the message (pattern with 3 groups)
                if len(match.groups()) >= 3 and match.group(3):
                    product_name_from_message = match.group(3).strip()
                
                logger.info(f"Direct add to cart: {quantity} x {product_code}")
                
                # Try direct processing first (bypass LLM for simple "add X CODE to cart" format)
                try:
                    # Get user and product
                    user = User.query.get(session_user_id)
                    if not user:
                        return jsonify({'error': 'User not found'}), 401
                    
                    # Find product - for MRs, check dealer stock first
                    product = None
                    product_name = product_name_from_message  # Use product name from message if provided
                    
                    if user.role == 'mr' and user.area:
                        # For MRs, get products from dealer stock in their area
                        dealer_products = db_service.get_products_from_dealer_stock(user.area)
                        for p in dealer_products:
                            # Handle dict from dealer stock
                            p_code = p.get('product_code', '') if isinstance(p, dict) else (p.product_code if hasattr(p, 'product_code') else '')
                            if p_code.upper() == product_code.upper():
                                # If product_name was provided in message, match by name too
                                if product_name:
                                    p_name = p.get('product_name', '') if isinstance(p, dict) else p.product_name
                                    if p_name.lower() == product_name.lower():
                                        # Found matching product in dealer stock
                                        product_name = p.get('product_name', '') if isinstance(p, dict) else p.product_name
                                        # Try to find Product record with matching product_id from dealer stock
                                        from app.models import Product
                                        product_id = p.get('product_id') if isinstance(p, dict) else (p.product_id if hasattr(p, 'product_id') else None)
                                        product = Product.query.get(product_id) if product_id else None
                                        if not product:
                                            # Try to find Product by name
                                            product = Product.query.filter_by(product_name=product_name).first()
                                        break
                                else:
                                    # No product name in message, use dealer stock name
                                    product_name = p.get('product_name') if isinstance(p, dict) else p.product_name
                                    # Try to find Product record
                                    from app.models import Product
                                    product_id = p.get('product_id') if isinstance(p, dict) else (p.product_id if hasattr(p, 'product_id') else None)
                                    product = Product.query.get(product_id) if product_id else None
                                    if not product:
                                        # Try to find Product by name
                                        product = Product.query.filter_by(product_name=product_name).first()
                                    break
                    
                    # Fallback to regular product lookup
                    if not product:
                        product = db_service.get_product_by_code(product_code)
                        if product and not product_name:
                            product_name = product.product_name
                    
                    if not product:
                        return jsonify({
                            'response': f'Product not found. Please check the product name and try again.',
                            'action_buttons': [
                                {'text': 'Try Again', 'action': 'place_order'}
                            ]
                        }), 200
                    
                    # Use product name from message/dealer stock if available, otherwise from product
                    if not product_name:
                        product_name = product.product_name
                    
                    # Calculate pricing
                    enhanced_order_service = get_enhanced_order_service()
                    pricing = enhanced_order_service.pricing_service.calculate_product_pricing(product.id, quantity)
                    
                    if 'error' in pricing:
                        return jsonify({
                            'response': f'Error calculating price for {product_name}: {pricing["error"]}',
                            'action_buttons': [
                                {'text': 'Try Again', 'action': 'place_order'}
                            ]
                        }), 200
                    
                    # Add to cart directly - pass product_name to distinguish products with same code
                    product_code = product_code or str(product.id)
                    unit_price = pricing.get('pricing', {}).get('final_price', 0)
                    cart_item, message = db_service.add_to_cart(
                        session_user_id,
                        product.id,
                        product_code,
                        product_name,
                        quantity,
                        unit_price
                    )
                    
                    # Block quantity for MR orders when adding to cart
                    if cart_item and user.role == 'mr' and user.area:
                        enhanced_order_service = get_enhanced_order_service()
                        warehouse = db_service.get_warehouse_by_area(user.area)
                        if warehouse:
                            enhanced_order_service._block_quantity_for_mr_order(
                                user=user,
                                product_code=product_code,
                                quantity=quantity,
                                warehouse=warehouse
                            )
                            logger.info(f"Blocked {quantity} units of {product_code} for MR order (added to cart)")
                    
                    if cart_item:
                        # Get updated cart items
                        cart_items = db_service.get_cart_items(session_user_id)
                        cart_data = []
                        for item in cart_items:
                            product_name = item.product_code
                            if hasattr(item, 'product') and item.product:
                                product_name = item.product.product_name
                            elif hasattr(item, 'product_name'):
                                product_name = item.product_name
                            
                            cart_data.append({
                                'id': item.id,
                                'product_code': item.product_code,
                                'product_name': product_name,
                                'quantity': item.quantity,
                                'unit_price': item.unit_price,
                                'total_price': item.total_price,
                                'base_price': getattr(item, 'base_price', item.unit_price),
                                'discount_amount': getattr(item, 'discount_amount', 0),
                                'final_price': getattr(item, 'final_price', item.unit_price),
                                'scheme_applied': getattr(item, 'scheme_applied', None),
                                'free_quantity': getattr(item, 'free_quantity', 0),
                                'paid_quantity': getattr(item, 'paid_quantity', item.quantity)
                            })
                        
                        save_conversation(user.id, user_message, f'Added {quantity} {product_name} to cart')
                        return jsonify({
                            'response': f'✅ Added {quantity} unit(s) of {product_name} to cart!',
                            'cart_items': cart_data,
                            'action_buttons': [
                                {'text': 'View Cart', 'action': 'view_cart'},
                                {'text': 'Add More Items', 'action': 'place_order'},
                                {'text': 'Place Order', 'action': 'confirm_order'}
                            ]
                        }), 200
                    else:
                        return jsonify({
                            'response': f'Failed to add {product_name} to cart: {message}',
                            'action_buttons': [
                                {'text': 'Try Again', 'action': 'place_order'},
                                {'text': 'View Cart', 'action': 'view_cart'}
                            ]
                        }), 200
                        
                except Exception as e:
                    logger.error(f"Error in direct add to cart: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    # Fallback to LLM-based processing
                    logger.info("Falling back to LLM-based processing")
                
                # This is an "add to cart" message - process it
                enhanced_order_service = get_enhanced_order_service()
                result = enhanced_order_service.process_order_request(user_message, session_user_id, conversation_history)
                
                if result.get('success'):
                    # Get updated cart items
                    cart_items = db_service.get_cart_items(session_user_id)
                    cart_data = []
                    for item in cart_items:
                        # Get product name from relationship or fallback
                        product_name = item.product_code
                        if hasattr(item, 'product') and item.product:
                            product_name = item.product.product_name
                        elif hasattr(item, 'product_name'):
                            product_name = item.product_name
                        
                        cart_data.append({
                            'id': item.id,
                            'product_code': item.product_code,
                            'product_name': product_name,
                            'quantity': item.quantity,
                            'unit_price': item.unit_price,
                            'total_price': item.total_price,
                            'base_price': getattr(item, 'base_price', item.unit_price),
                            'discount_amount': getattr(item, 'discount_amount', 0),
                            'final_price': getattr(item, 'final_price', item.unit_price),
                            'scheme_applied': getattr(item, 'scheme_applied', None),
                            'free_quantity': getattr(item, 'free_quantity', 0),
                            'paid_quantity': getattr(item, 'paid_quantity', item.quantity)
                        })
                    
                    save_conversation(user.id, user_message, result.get('message', 'Items added to cart'))
                    return jsonify({
                        'response': result.get('message', 'Items added to cart'),
                        'cart_items': cart_data,
                        'action_buttons': [
                            {'text': 'View Cart', 'action': 'view_cart'},
                            {'text': 'Add More Items', 'action': 'place_order'},
                            {'text': 'Place Order', 'action': 'confirm_order'}
                        ]
                    }), 200
                else:
                    save_conversation(user.id, user_message, result.get('message', 'Error adding items to cart'))
                    return jsonify({
                        'response': result.get('message', 'Error adding items to cart'),
                        'action_buttons': [
                            {'text': 'Try Again', 'action': 'place_order'},
                            {'text': 'View Cart', 'action': 'view_cart'}
                        ]
                    }), 200
        
        # Ensure context_data is always defined before classification
        if 'context_data' not in locals() or context_data is None:
            context_data = {
                'user_warehouse': area,
                'user_email': user.email if user else None,
                'user_type': user.role if user else None,
                'user_role': user.role if user else None
            }
        
        # Classify user intent using LLM (only if not stock confirmation and not add to cart)
        try:
            classification_result = classification_service.classify_user_intent(user_message, context_data)
        except Exception as e:
            logger.error(f"Error in classification service: {str(e)}", exc_info=True)
            # Fallback to default classification
            classification_result = {'classification': 'OTHER', 'confidence': 0.0}
        intent = classification_result.get('classification', 'OTHER')
        
        logger.info(f"Intent classified as: {intent}")
        
        # Process based on classification
        
        # Check for action buttons first
        message_lower = user_message.lower().strip()
        
        # Process based on classification and add intent to response
        response = None
        if intent == 'PLACE_ORDER':
            response = handle_place_order(user_message, user, context_data, conversation_history)
        elif intent == 'TRACK_ORDER':
            response = handle_track_order(user_message, user, context_data)
        elif intent == 'CALCULATE_COST':
            response = handle_calculate_cost(user_message, user, context_data, conversation_history)
        elif intent == 'COMPANY_INFO':
            response = handle_company_info(user_message, user)
        elif intent == 'WEB_SEARCH':
            response = handle_web_search(user_message, user, context_data)
        elif intent == 'PRODUCT_INFO':
            response = handle_product_info_or_query(user_message, user, context_data)
        else:
            response = handle_general_conversation(user_message, user, context_data)
        
        # Intent tracking (voice disabled)
        if response:
            # Extract response data if it's a tuple (jsonify response)
            if isinstance(response, tuple):
                response_obj, status_code = response
                if hasattr(response_obj, 'get_json'):
                    data = response_obj.get_json()
                    if data:
                        data['intent'] = intent
                        # Translate response before returning
                        data = translate_response(data, user_language)
                        return jsonify(data), status_code
            elif isinstance(response, dict):
                response['intent'] = intent
                return response
        
        return response

    except Exception as e:
        logger.error(f"Error processing message: {str(e)}")
        user = User.query.get(session.get('user_id')) if session.get('user_id') else None
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error. Please try again.'}), user), 500

def handle_order_confirmation(user, user_id):
    """Handle order confirmation and placement"""
    try:
        enhanced_order_service = get_enhanced_order_service()
        
        # Place order from cart
        # Get customer details if MR - customer selection is mandatory for MRs
        customer_id = None
        if user.role == 'mr':
            if 'selected_customer_id' not in session:
                response = "To place an order, I need to know which customer you're ordering for. Please select a customer first by clicking 'Place Order'."
                save_conversation(user.id, "confirm order", response)
                return jsonify({
                    'response': response,
                    'requires_customer_selection': True,
                    'action_buttons': [
                        {'text': 'Place Order', 'action': 'place_order'},
                        {'text': 'View Open Order', 'action': 'open_order'},
                        {'text': 'Company Info', 'action': 'company_info'},
                        {'text': 'Product Info', 'action': 'product_info'}
                    ]
                }), 200
            customer_id = session.get('selected_customer_id')
        
        result = enhanced_order_service.place_order(user_id, customer_id=customer_id)
        
        if result['success']:
            save_conversation(user.id, "confirm order", result['message'])
            
            return jsonify({
                'response': result['message'],
                'order_id': result.get('order_id'),
                'order_summary': result.get('order_summary', {}),
                'next_steps': result.get('next_steps', ''),
                'action_buttons': [
                    {'text': 'Track Order', 'action': 'track_order'},
                    {'text': 'Place New Order', 'action': 'place_order'}
                ]
            }), 200
        else:
            save_conversation(user.id, "confirm order", result['message'])
            return jsonify({
                'response': result['message'],
                'action_buttons': [
                    {'text': 'View Cart', 'action': 'view_cart'},
                    {'text': 'Get Help', 'action': 'help'}
                ]
            }), 200
            
    except Exception as e:
        logger.error(f"Error handling order confirmation: {str(e)}")
        user = User.query.get(user_id) if user_id else None
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error placing your order. Please try again.'}), user), 500

def handle_place_order(user_message, user, context_data, conversation_history):
    """Handle order placement requests - Show customer selection for MRs, then product selection"""
    try:
        db_service = get_db_service()
        
        logger.info(f"handle_place_order called. User type: {user.role}, User ID: {user.id}")
        logger.info(f"Session keys: {list(session.keys())}")
        logger.info(f"selected_customer_id in session: {'selected_customer_id' in session}")
        
        # For MRs, check if customer is already selected
        if user.role == 'mr':
            # Check if customer is already selected in session
            if 'selected_customer_id' not in session:
                # First, ask whether to select existing customer or add new customer
                response = "To place an order, I need to know which customer you're ordering for. Would you like to select an existing customer or add a new customer?"
                
                logger.info(f"MR without selected customer. Returning customer selection buttons.")
                save_conversation(user.id, user_message, response)
                
                return jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'Select Customer', 'action': 'select_customer'},
                        {'text': 'Add New Customer', 'action': 'add_new_customer'}
                    ]
                }), 200
            
            # Customer is selected, proceed with product selection
            selected_customer = Customer.query.get(session.get('selected_customer_id'))
            if not selected_customer:
                # Customer not found, reset selection
                session.pop('selected_customer_id', None)
                return handle_place_order(user_message, user, context_data, conversation_history)
        
        # For non-MRs, proceed directly to product selection
        
        # Get available products for the user
        if user.role == 'mr' and user.area:
            # Get products from dealer stock in MR's area
            products = db_service.get_products_from_dealer_stock(user.area)
        elif user.role == 'distributor' and user.area:
            # For distributors, get products from dealer_wise_stock_details (their own stock)
            products = db_service.get_products_from_dealer_stock(user.area)
        else:
            # Fallback: use Product table only if no area or not MR/distributor
            warehouse = db_service.get_warehouse_by_area(user.area) if user.area else None
            if warehouse:
                products = db_service.get_products_by_warehouse(warehouse.id)
            else:
                from app.models import Product
                products = Product.query.all()
        
        if not products:
            response = "No products are currently available in your area. Please contact support."
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Get Help', 'action': 'help'}
                ]
            }), 200
        
        # Build product list for interactive UI with FOC information using helper
        pricing_service = get_pricing_service()
        product_list = build_product_list_with_foc(products, pricing_service)
        
        # Show customer info if MR
        customer_info = ""
        if user.role == 'mr' and 'selected_customer_id' in session:
            selected_customer = Customer.query.get(session.get('selected_customer_id'))
            if selected_customer:
                customer_info = f"• **Ordering for:** {selected_customer.name} ({selected_customer.unique_id})\n"
        
        response = f"• Great! I can help you place an order.\n{customer_info}• Please use the product selection form below to select products and quantities."
        
        save_conversation(user.id, user_message, response)
        
        # Pass change customer info to frontend so it can add button inside the form
        # But don't show action buttons at the bottom
        action_buttons = []
        if user.role == 'mr' and 'selected_customer_id' in session:
            # Pass this info so frontend can add change customer button in the form
            action_buttons.append({'text': 'Change Customer', 'action': 'change_customer'})
        
        return jsonify({
            'response': response,
            'products': product_list,
            'interactive_product_selection': True,  # Flag for frontend to render interactive UI
            'show_product_table': True,  # Flag to show table outside selection box
            'action_buttons': action_buttons  # Pass info but don't display - buttons are in the form
        }), 200
            
    except Exception as e:
        logger.error(f"Error handling place order: {str(e)}")
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error processing your order. Please try again.'}), user), 500

def handle_track_order(user_message, user, context_data=None):
    """Handle track order requests. context_data is optional."""
    if context_data is None:
        context_data = {}
    db_service = get_db_service()
    enhanced_order_service = get_enhanced_order_service()
    try:
        import re
        from app.models import Order
        from datetime import datetime, timedelta
        
        message_lower = user_message.lower()
        order_id_pat = r'([A-Z]{2}\d{8}\w+)'  # e.g. QB20251030F475CA7D
        match_id = re.search(order_id_pat, user_message)
        order_id = match_id.group(1) if match_id else None
        
        # Check if this is a confirmation, rejection, or cancellation request
        is_confirm_request = any(keyword in message_lower for keyword in ['confirm order', 'confirm the order']) and order_id
        is_reject_request = any(keyword in message_lower for keyword in ['reject order', 'reject the order']) and order_id and user.role == 'distributor'
        is_cancel_request = any(keyword in message_lower for keyword in ['cancel order', 'cancel my order', 'cancel the order']) and order_id and user.role == 'mr'
        
        # If MR and cancel intent
        if user.role == 'mr' and order_id and is_cancel_request:
            # Attempt to cancel order
            cancel_res = enhanced_order_service.cancel_order_by_mr(order_id, user.id)
            if cancel_res['success']:
                response_msg = f"❌ **Order Cancelled**\n\n{cancel_res['message']}\n\nThe order has been cancelled and stock has been unblocked."
                action_buttons = [
                    {'text': 'View All Orders', 'action': 'track_order'},
                    {'text': 'Place New Order', 'action': 'place_order'},
                    {'text': 'Back to Home', 'action': 'home'}
                ]
                return jsonify({'response': response_msg, 'action_buttons': action_buttons}), 200
            else:
                return ensure_action_buttons(jsonify({'response': cancel_res['message']}), user), 200
        
        # If distributor and reject intent
        if user.role == 'distributor' and order_id and is_reject_request:
            # Attempt to reject
            reject_res = enhanced_order_service.reject_order_by_distributor(order_id, user.id)
            if reject_res['success']:
                response_msg = f"❌ **Order Rejected**\n\n{reject_res['message']}\n\nThe stock has been unblocked and is now available for other orders."
                action_buttons = [
                    {'text': 'View All Orders', 'action': 'track_order'},
                    {'text': 'Back to Home', 'action': 'home'}
                ]
                return jsonify({'response': response_msg, 'action_buttons': action_buttons}), 200
            else:
                return ensure_action_buttons(jsonify({'response': reject_res['message']}), user), 200
        
        # If distributor and confirm intent
        if user.role == 'distributor' and order_id and is_confirm_request:
            # Attempt to confirm:
            confirm_res = enhanced_order_service.confirm_order_by_distributor(order_id, user.id)
            if confirm_res['success']:
                # After confirmation, present the order table with updated status
                stat = enhanced_order_service.get_order_status_for_distributor(order_id, user.id)
                response_msg = stat['message'] + "\n\n✅ **Order successfully confirmed!**"
                action_buttons = [
                    {'text': 'Track Another Order', 'action': 'track_order'},
                    {'text': 'View All Orders', 'action': 'track_order'},
                    {'text': 'Place New Order', 'action': 'place_order'}
                ]
                return jsonify({'response': response_msg, 'action_buttons': action_buttons}), 200
            else:
                return ensure_action_buttons(jsonify({'response': confirm_res['message']}), user), 200
        
        # Otherwise, follow normal track logic
        if user.role == 'distributor':
            if order_id:
                # Show specific order details
                order_stat = enhanced_order_service.get_order_status_for_distributor(order_id, user.id)
                if order_stat['success']:
                    order = Order.query.filter_by(order_id=order_id).first()
                    if order:
                        # Get order items with FOC - use adjusted quantities if available
                        items_list = []
                        total_items = 0
                        for item in order.order_items:
                            # Use adjusted_quantity if available (after dealer confirmation), otherwise use original quantity
                            actual_quantity = item.adjusted_quantity if item.adjusted_quantity is not None else item.quantity
                            free_qty = item.free_quantity or 0
                            total_qty = actual_quantity + free_qty
                            total_items += total_qty
                            
                            # Recalculate total_price based on adjusted quantity if quantity was adjusted
                            unit_price = float(item.unit_price) if item.unit_price else 0.0
                            if item.adjusted_quantity is not None and item.adjusted_quantity != item.quantity:
                                # Quantity was adjusted, recalculate total_price based on adjusted quantity
                                total_price = float(actual_quantity * unit_price)
                            else:
                                # Use original total_price if quantity wasn't adjusted
                                total_price = float(item.total_price) if item.total_price else 0.0
                            
                            items_list.append({
                                'product_name': item.product.product_name,
                                'product_code': item.product_code,
                                'quantity': actual_quantity,  # Use adjusted quantity if available
                                'free_quantity': free_qty,
                                'total_quantity': total_qty,
                                'unit_price': unit_price,
                                'total_price': total_price  # Recalculated if quantity was adjusted
                            })
                        
                        # Check if this order was created by fulfilling a pending order
                        fulfilled_pending = PendingOrderProducts.query.filter_by(fulfilled_order_id=order.order_id).all()
                        
                        # Build response with order details using bullet points
                        msg = f"**📦 Order Details**\n\n"
                        
                        # Add special note for fulfilled pending orders
                        if fulfilled_pending:
                            msg += "⚠️ **This order was created by fulfilling a previous pending order.**\n"
                            for p in fulfilled_pending:
                                original_order = None
                                if p.original_order_id:
                                    original_order = Order.query.filter_by(order_id=p.original_order_id).first()
                                
                                original_date = original_order.created_at.strftime('%B %d, %Y at %I:%M %p') if (original_order and original_order.created_at) else 'N/A'
                                msg += f"• Original Order ID: **{p.original_order_id or 'N/A'}** (placed on {original_date})\n"
                                msg += f"  - Product: {p.product_name} ({p.product_code})\n"
                                msg += f"  - Requested Quantity: {p.requested_quantity} units\n"
                            msg += "\n"
                        
                        # Add concise timeline information
                        msg += "**⏱️ Order Timeline:**\n"
                        placed_at = order.created_at.strftime('%B %d, %Y at %I:%M %p') if order.created_at else 'N/A'
                        confirmed_at = order.distributor_confirmed_at.strftime('%B %d, %Y at %I:%M %p') if order.distributor_confirmed_at else None
                        delivered_at = order.delivered_at.strftime('%B %d, %Y at %I:%M %p') if order.delivered_at else None
                        
                        msg += f"• Order placed: {placed_at}\n"
                        msg += f"• Dealer confirmed: {confirmed_at or 'Not yet confirmed by dealer'}\n"
                        msg += f"• Delivered to customer: {delivered_at or 'Not yet delivered'}\n\n"
                        
                        # Standard order information block
                        msg += f"**📋 Order Information:**\n"
                        msg += f"• **Order ID:** {order.order_id}\n"
                        msg += f"• **Status:** {order.status.replace('_', ' ').title()}\n"
                        msg += f"• **MR:** {order.mr.name if order.mr else 'N/A'}\n"
                        msg += f"• **Area:** {order.mr.area if order.mr else 'N/A'}\n"
                        msg += f"• **Date:** {order.created_at.strftime('%B %d, %Y at %I:%M %p') if order.created_at else 'N/A'}\n"
                        msg += f"• **Total Items:** {total_items} units\n\n"
                        
                        msg += "**🛍️ Order Items:**\n"
                        for item in items_list:
                            if item['free_quantity'] > 0:
                                msg += f"• **{item['product_name']}** ({item['product_code']})\n"
                                msg += f"  - Quantity: {item['quantity']} paid + **{item['free_quantity']} FREE** = **{item['total_quantity']} total units**\n"
                                msg += f"  - Unit Price: {item['unit_price']:,.2f} MMK\n"
                                msg += f"  - Total: {item['total_price']:,.2f} MMK\n"
                            else:
                                msg += f"• **{item['product_name']}** ({item['product_code']})\n"
                                msg += f"  - Quantity: {item['quantity']} units\n"
                                msg += f"  - Unit Price: {item['unit_price']:,.2f} MMK\n"
                                msg += f"  - Total: {item['total_price']:,.2f} MMK\n"
                        
                        # Add tax information
                        if hasattr(order, 'subtotal') and order.subtotal:
                            msg += f"\n**💰 Payment Summary:**\n"
                            msg += f"• **Subtotal:** {order.subtotal:,.2f} MMK\n"
                            msg += f"• **Tax (5%):** {order.tax_amount:,.2f} MMK\n"
                            msg += f"• **Grand Total:** {order.total_amount:,.2f} MMK\n"
                        else:
                            msg += f"\n**💰 Total Amount:** {order.total_amount:,.2f} MMK\n"
                        
                        # Add action buttons based on order status and user role
                        actions = []
                        if user.role == 'distributor':
                            if order.status == 'pending':
                                actions.append({'text': '✅ Confirm Order', 'action': f'confirm_order_{order_id}'})
                                actions.append({'text': '❌ Reject Order', 'action': f'reject_order_{order_id}'})
                        elif user.role == 'mr':
                            if order.status in ['pending', 'draft']:
                                actions.append({'text': '❌ Cancel Order', 'action': f'cancel_order_{order_id}'})
                        actions.append({'text': 'View All Orders', 'action': 'track_order'})
                        actions.append({'text': 'Back to Home', 'action': 'home'})
                        
                        return jsonify({'response': msg, 'action_buttons': actions}), 200
                    else:
                        return ensure_action_buttons(jsonify({'response': 'Order not found'}), user), 200
                else:
                    return ensure_action_buttons(jsonify({'response': order_stat['message']}), user), 200
            else:
                # Check if user is asking for specific status
                message_lower = user_message.lower()
                status_filter = None
                filter_description = "all orders"
                
                # Detect status filter in message
                # Note: "pending orders" (out-of-stock orders) is different from "pending stocks" (stock arrivals to confirm)
                # This handler only shows out-of-stock orders, not stock arrivals
                if 'pending' in message_lower and 'stock' not in message_lower:
                    # Special handling for pending orders - show PendingOrderProducts (out-of-stock orders)
                    pending_items = db_service.get_pending_order_products(area=user.area, status='pending')
                    if not pending_items:
                        return ensure_action_buttons(jsonify({'response': 'No pending orders found in your warehouse.'}), user), 200
                    
                    summary = "**Pending orders for your warehouse:**\n\n"
                    summary += "| Product Code | Product Name | Requested Qty | Customer | Order ID | Requested Date |\n"
                    summary += "|--------------|--------------|---------------|----------|----------|----------------|\n"
                    for item in pending_items:
                        customer = User.query.get(item.user_id)
                        customer_name = customer.name if customer else 'Unknown'
                        order_ref = item.original_order_id if item.original_order_id else 'N/A'
                        summary += f"| {item.product_code} | {item.product_name} | {item.requested_quantity} | {customer_name} | {order_ref} | {item.created_at.strftime('%Y-%m-%d')} |\n"
                    summary += '\n**Note:** These products are waiting for stock to arrive. They will be automatically ordered when available.'
                    return ensure_action_buttons(jsonify({'response': summary}), user), 200
                elif 'confirmed' in message_lower:
                    status_filter = ['confirmed', 'distributor_confirmed']
                    filter_description = "confirmed orders"
                elif 'in transit' in message_lower or 'transit' in message_lower:
                    status_filter = ['in_transit', 'distributor_notified']
                    filter_description = "in-transit orders"
                elif 'shipped' in message_lower:
                    status_filter = ['shipped']
                    filter_description = "shipped orders"
                elif 'delivered' in message_lower:
                    status_filter = ['delivered']
                    filter_description = "delivered orders"
                
                # Get all orders for distributor's area
                # Get orders from distributor's area
                orders = Order.query.join(User, Order.mr_id == User.id).filter(
                    User.area == user.area
                ).order_by(Order.created_at.desc()).all()
                
                if not orders:
                    return ensure_action_buttons(jsonify({'response': 'No orders found in your area.'}), user), 200
                
                # Get unique MRs, statuses, and dates for filters
                unique_mrs = sorted(list(set([o.mr.name for o in orders if o.mr])))
                unique_statuses = sorted(list(set([o.status for o in orders if o.status])))
                unique_dates = sorted(list(set([o.created_at.strftime('%Y-%m-%d') for o in orders if o.created_at])), reverse=True)
                
                # Build orders list with details
                orders_data = []
                for o in orders[:50]:  # Show latest 50 orders
                    orders_data.append({
                        'order_id': o.order_id,
                        'mr_name': o.mr.name if o.mr else 'N/A',
                        'mr_id': o.mr_unique_id if o.mr_unique_id else 'N/A',
                        'status': o.status,
                        'status_display': o.status.replace('_', ' ').title() if o.status else 'Unknown',
                        'total_amount': float(o.total_amount) if o.total_amount else 0.0,
                        'order_date': o.created_at.strftime('%Y-%m-%d') if o.created_at else 'N/A',
                        'order_time': o.created_at.strftime('%H:%M') if o.created_at else 'N/A',
                        'order_datetime': o.created_at.strftime('%Y-%m-%d %H:%M') if o.created_at else 'N/A',
                        'area': o.mr.area if o.mr else 'N/A',
                        'can_confirm': o.status == 'pending',
                        'items_count': o.order_items.count() if hasattr(o.order_items, 'count') else 0
                    })
                
                # Return selection box for distributors
                # Build response in English first (for database)
                response_msg_en = f"**📊 Order Management - {user.area} Area**\n\n"
                response_msg_en += f"Found **{len(orders_data)}** orders in your area.\n\n"
                response_msg_en += "**Use filters below to narrow down orders:**\n"
                response_msg_en += "• Filter by MR Name\n"
                response_msg_en += "• Filter by Status (Pending, Confirmed, Rejected)\n"
                response_msg_en += "• Filter by Date\n\n"
                response_msg_en += "Select an order to view details and take action (Confirm/Reject)."
                
                # Save English version to database
                save_conversation(user.id, user_message, response_msg_en)
                
                # Translate for user if needed
                user_language = session.get('user_language', 'en')
                response_msg = response_msg_en
                if user_language != 'en':
                    translation_service = get_translation_service()
                    if translation_service.is_available():
                        response_msg = translation_service.translate(response_msg_en, user_language)
                
                return ensure_action_buttons(jsonify({
                    'response': response_msg,
                    'interactive_order_selection': True,  # Flag for frontend to show selection box
                    'order_selection_type': 'distributor',  # Type of selection
                    'orders': orders_data,
                    'show_orders_table': True,  # Show orders in table format first
                    'filters': {
                        'mr_names': unique_mrs,
                        'statuses': unique_statuses,
                        'dates': unique_dates
                    }
                }), user), 200
        # fall back to self-tracking for non-distributors
        if order_id:
            status = enhanced_order_service.get_order_status(order_id, user.id)
            if status['success']:
                # Build order details table
                table = '| Product | Quantity | Unit Price | Total |\n|--------|----------|-----------|-------|\n'
                for i in status['order']['items']:
                    # Get quantity (handle different field names)
                    quantity = i.get('quantity') or i.get('product_quantity_ordered') or 0
                    # Get free quantity if available
                    free_quantity = i.get('free_quantity', 0)
                    # Format quantity display
                    if free_quantity and free_quantity > 0:
                        qty_display = f"{quantity} + {free_quantity} free = {quantity + free_quantity}"
                    else:
                        qty_display = str(quantity) if quantity else '0'
                    
                    # Build table row with safe field access
                    product_name = i.get('product_name') or i.get('product_code', 'Unknown')
                    product_code = i.get('product_code', 'N/A')
                    unit_price = float(i.get('unit_price', 0)) if i.get('unit_price') else 0.0
                    total_price = float(i.get('total_price', 0)) if i.get('total_price') else 0.0
                    
                    table += f"| {product_name} | {qty_display} | {unit_price:,.2f} MMK | {total_price:,.2f} MMK |\n"
                
                # Format status display
                status_display = status['order']['status'].replace('_', ' ').title()
                order_stage_display = status['order'].get('order_stage', '').replace('_', ' ').title() if status['order'].get('order_stage') else ''
                
                # Build message with tax information
                msg = f"**📦 Track Order - {order_id}:**\n\n"
                msg += f"**📊 Order Information:**\n"
                msg += f"• **Status:** {status_display}\n"
                if order_stage_display:
                    msg += f"• **Order Stage:** {order_stage_display}\n"
                if status['order'].get('order_date'):
                    from datetime import datetime
                    try:
                        order_date = datetime.fromisoformat(status['order']['order_date'].replace('Z', '+00:00'))
                        msg += f"• **Order Date:** {order_date.strftime('%B %d, %Y at %I:%M %p')}\n"
                    except:
                        msg += f"• **Order Date:** {status['order']['order_date']}\n"
                
                msg += f"\n{table}\n"
                
                # Get order object to show tax breakdown
                order_obj = Order.query.filter_by(order_id=order_id).first()
                if order_obj and hasattr(order_obj, 'subtotal') and order_obj.subtotal:
                    msg += f"\n**💰 Payment Summary:**\n"
                    msg += f"• **Subtotal:** ${order_obj.subtotal:,.2f}\n"
                    msg += f"• **Tax (5%):** ${order_obj.tax_amount:,.2f}\n"
                    msg += f"• **Grand Total:** ${order_obj.total_amount:,.2f}\n"
                else:
                    # Old orders without tax - show total only
                    msg += f"\n**💰 Total Amount:** ${status['order']['total_amount']:,.2f}\n"
                
                # Add action buttons to continue conversation
                action_buttons = [
                    {'text': 'Track Another Order', 'action': 'track_order'},
                    {'text': 'Place New Order', 'action': 'place_order'}
                ]
                
                return jsonify({
                    'response': msg,
                    'action_buttons': action_buttons,
                    'animate_order_details': True  # Flag to trigger staggered animation
                }), 200
            else:
                return ensure_action_buttons(jsonify({'response':status['message']}), user), 200
        # Get orders by mr_id (for MRs)
        if user.role == 'mr':
            # For MRs, get orders where they are the MR
            orders = Order.query.filter(
                Order.mr_id == user.id
            ).order_by(Order.created_at.desc()).all()
            
            # Get unique customers, statuses, and dates for filters
            unique_customers = []
            unique_statuses = sorted(list(set([o.status for o in orders if o.status])))
            unique_dates = sorted(list(set([o.created_at.strftime('%Y-%m-%d') for o in orders if o.created_at])), reverse=True)
            
            # Get unique customers from orders
            customer_ids = set([o.customer_id for o in orders if o.customer_id])
            from app.models import Customer
            customers = Customer.query.filter(Customer.id.in_(customer_ids)).all() if customer_ids else []
            unique_customers = sorted([f"{c.name} ({c.unique_id})" for c in customers])
            
            # Prepare orders data for frontend display with customer info
            orders_list = []
            for o in orders[:50]:  # Show up to 50 recent orders
                status_display = (o.status or o.order_stage or 'Unknown').replace('_', ' ').title()
                customer_name = 'N/A'
                customer_id = None
                if o.customer:
                    customer_name = f"{o.customer.name} ({o.customer.unique_id})"
                    customer_id = o.customer.unique_id
                
                orders_list.append({
                    'order_id': o.order_id,
                    'status': status_display,
                    'total_amount': float(o.total_amount) if o.total_amount else 0.0,
                    'order_date': o.created_at.strftime('%Y-%m-%d') if o.created_at else 'N/A',
                    'status_raw': o.status or o.order_stage or 'unknown',
                    'customer_name': customer_name,
                    'customer_id': customer_id
                })
            
            # Return response with orders data and filters for interactive display (NO MR filter for MR users)
            # Build response in English first (for database)
            response_msg_en = f"**📊 Your Orders**\n\n"
            response_msg_en += f"Found **{len(orders)}** order(s) in total.\n\n"
            response_msg_en += "**Use the filters below to narrow down your search:**\n"
            response_msg_en += "• Filter by Customer\n"
            response_msg_en += "• Filter by Status\n"
            response_msg_en += "• Filter by Date\n\n"
            response_msg_en += "**Select an order from the dropdown to view details.**"
            
            # Save English version to database
            save_conversation(user.id, user_message, response_msg_en)
            
            # Translate for user if needed
            user_language = session.get('user_language', 'en')
            response_msg = response_msg_en
            if user_language != 'en':
                translation_service = get_translation_service()
                if translation_service.is_available():
                    response_msg = translation_service.translate(response_msg_en, user_language)
            
            return jsonify({
                'response': response_msg,
                'show_orders_table': True,
                'orders': orders_list,
                'interactive_order_selection': True,
                'order_selection_type': 'mr',  # Type of selection
                'filters': {
                    'customers': unique_customers,
                    'statuses': unique_statuses,
                    'dates': unique_dates
                    # Note: No 'mr_names' filter for MR users - they only see their own orders
                },
                'action_buttons': []  # Prevent showing action buttons when order selection is active
            }), 200
        else:
            # For other users, get orders by user_id
            orders = db_service.get_orders_by_user(user.id)
        
        if not orders:
            return ensure_action_buttons(jsonify({'response':'No orders found in your account.'}), user), 200
        
        # Prepare orders data for frontend display
        orders_list = []
        for o in orders[:10]:  # Show up to 10 recent orders
            status_display = (o.status or o.order_stage or 'Unknown').replace('_', ' ').title()
            orders_list.append({
                'order_id': o.order_id,
                'status': status_display,
                'total_amount': float(o.total_amount) if o.total_amount else 0.0,
                'order_date': o.created_at.strftime('%Y-%m-%d') if o.created_at else 'N/A',
                'status_raw': o.status or o.order_stage or 'unknown'
            })
        
        # Return response with orders data for interactive display
        # Build response in English first
        response_text_en = 'Here are your recent orders. Select an order to view details:'
        
        # Translate for user if needed
        user_language = session.get('user_language', 'en')
        response_text = response_text_en
        if user_language != 'en':
            translation_service = get_translation_service()
            if translation_service.is_available():
                response_text = translation_service.translate(response_text_en, user_language)
        
        # Save English version to database
        save_conversation(user.id, user_message, response_text_en)
        
        return ensure_action_buttons(jsonify({
            'response': response_text,
            'show_orders_table': True,
            'orders': orders_list,
            'interactive_order_selection': True,
            'action_buttons': []  # Prevent showing action buttons when order selection is active
        }), user), 200
    except Exception as e:
        import traceback
        traceback.print_exc()
        logger.error(f"Error handling track order: {str(e)}")
        return ensure_action_buttons(jsonify({'response':'Sorry, I encountered an error tracking your order. Please try again.'}), user), 500

def handle_calculate_cost(user_message, user, context_data, conversation_history):
    """Handle cost calculation requests"""
    try:
        # Get cart items
        cart_items = db_service.get_cart_items(user.id)
        
        if not cart_items:
            response = "Your cart is empty. Would you like to add some products first?"
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Add Products', 'action': 'place_order'}
                ]
            }), 200
        
        # Generate order summary with pricing
        llm_order_service = get_llm_order_service()
        order_summary = llm_order_service.generate_order_summary(cart_items, user)
        
        response = order_summary['summary']
        save_conversation(user.id, user_message, response)
        
        return jsonify({
            'response': response,
            'order_summary': order_summary,
            'action_buttons': [
                {'text': 'Place Order', 'action': 'place_order'},
                {'text': 'Modify Cart', 'action': 'view_cart'},
                {'text': 'Add More Items', 'action': 'add_items'}
            ]
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling calculate cost: {str(e)}")
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error calculating the cost. Please try again.'}), user), 500

def handle_company_info(user_message, user):
    """Handle company information requests"""
    try:
        # Use default company info for Quantum Blue AI
        company_info = {
            'company_name': 'Quantum Blue AI',
            'description': 'Your intelligent assistant for orders, tracking, and more!',
            'features': [
                'Order Management',
                'Order Tracking',
                'Product Information',
                'Customer Management',
                'Stock Management'
            ],
            'contact_info': {
                'email': 'info@quantumblue.ai',
                'phone': '+1-800-QUANTUM',
                'address': 'Quantum Blue AI Headquarters'
            }
        }
        
        response = f"Welcome to {company_info['company_name']}!\n\n"
        response += f"{company_info['description']}\n\n"
        response += "Our features include:\n"
        for feature in company_info['features']:
            response += f"• {feature}\n"
        
        response += f"\nContact Information:\n"
        response += f"Email: {company_info['contact_info']['email']}\n"
        response += f"Phone: {company_info['contact_info']['phone']}\n"
        response += f"Address: {company_info['contact_info']['address']}\n"
        
        save_conversation(user.id, user_message, response)
        return jsonify({
            'response': response,
            'action_buttons': [
                {'text': 'Place Order', 'action': 'place_order'},
                {'text': 'View Open Order', 'action': 'open_order'},
                {'text': 'Company Info', 'action': 'company_info'}
            ]
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling company info: {str(e)}")
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error retrieving company information.'}), user), 500

def handle_product_info_or_query(user_message, user, context_data):
    """Handle product information and database queries"""
    # Get user language at the start of the function
    try:
        user_language = session.get('user_language', 'en')
    except:
        user_language = 'en'
    
    # Check if this is specifically the Product Info button click
    message_lower = user_message.lower().strip()
    if message_lower in ['product info', 'productinfo']:
        try:
            # Return product search interface
            search_service = get_search_service()
            if not search_service or not search_service.is_available():
                logger.warning("Azure AI Search not available for Product Info")
                response = "Product Info feature requires Azure AI Search to be configured. Please contact your administrator."
                save_conversation(user.id, user_message, response)
                return jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'Place Order', 'action': 'place_order'},
                        {'text': 'View Open Order', 'action': 'open_order'},
                        {'text': 'Company Info', 'action': 'company_info'}
                    ]
                }), 200
            
            # Get all available products for the selection list
            try:
                products = search_service.get_all_products(top=100)
                product_list = []
                for product in products:
                    # Extract product name from various possible fields
                    product_name = product.get('product_name') or product.get('name') or product.get('title') or 'Unknown Product'
                    product_list.append({
                        'id': product.get('id') or product.get('product_id') or product_name,
                        'name': product_name,
                        'description': product.get('description') or product.get('content') or product.get('text') or ''
                    })
                
                logger.info(f"Product Info: Retrieved {len(product_list)} products for user {user.id}")
            except Exception as e:
                logger.error(f"Error retrieving products for Product Info: {str(e)}")
                product_list = []
            
            save_conversation(user.id, user_message, "Product Info search interface opened")
            response_text = '🔍 **Product Information Search**\n\nPlease use the search interface below to find product information. You can search by product name or browse the available products.'
            # Translate response (user_language already defined at function start)
            if user_language != 'en':
                translation_service = get_translation_service()
                if translation_service.is_available():
                    response_text = translation_service.translate(response_text, user_language)
            
            return ensure_action_buttons(jsonify({
                'response': response_text,
                'interactive_product_search': True,
                'products': product_list,
                'action_buttons': [
                    {'text': 'Place Order', 'action': 'place_order'},
                    {'text': 'View Open Order', 'action': 'open_order'},
                    {'text': 'Company Info', 'action': 'company_info'}
                ]
            }), user), 200
        except Exception as e:
            logger.error(f"Error retrieving products from Azure Search: {str(e)}")
            response = f"Product Info feature is temporarily unavailable. Error: {str(e)}"
            # Translate error message (user_language already defined at function start)
            if user_language != 'en':
                translation_service = get_translation_service()
                if translation_service.is_available():
                    response = translation_service.translate(response, user_language)
            save_conversation(user.id, user_message, response)
            return ensure_action_buttons(jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'Place Order', 'action': 'place_order'},
                        {'text': 'View Open Order', 'action': 'open_order'},
                        {'text': 'Company Info', 'action': 'company_info'}
                    ]
                }), user), 200
        except Exception as e:
            logger.error(f"Error in handle_product_info_or_query for Product Info button: {str(e)}")
            response = "Sorry, I encountered an error while opening the Product Info feature. Please try again."
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Place Order', 'action': 'place_order'},
                    {'text': 'View Open Order', 'action': 'open_order'},
                    {'text': 'Company Info', 'action': 'company_info'}
                ]
            }), 200
    
    try:
        db_service = get_db_service()
        llm_service = get_llm_service()
        
        # For distributors, double-check if they meant "pending stock" not "products"
        if user.role == 'distributor':
            message_lower = user_message.lower().strip()
            # If message contains "pending" or "stock" keywords, redirect to stock confirmation
            if any(kw in message_lower for kw in ['pending', 'stock', 'arrival']):
                logger.info(f"Redirecting distributor product query to stock confirmation: {user_message}")
                return handle_stock_confirmation(user_message, user, context_data)
            
            # Check if this is a complex database query
            complex_query_indicators = [
                'how many', 'total', 'count', 'sum', 'average', 'top', 'most', 
                'least', 'analyze', 'statistics', 'stats', 'report', 'summary',
                'revenue', 'sales', 'quantity', 'amount'
            ]
            
            is_complex_query = any(indicator in message_lower for indicator in complex_query_indicators)
            
            if is_complex_query:
                # Use LLM to generate SQL or provide analytics
                return handle_distributor_analytics(user_message, user, db_service, llm_service)
        
        # For all users, handle simple product listing
        # For MRs, get products from dealer_wise_stock_details; for distributors, get from their dealer stock too
        if user.role == 'mr' and user.area:
            # Get products from dealer stock in MR's area
            products = db_service.get_products_from_dealer_stock(user.area)
        elif user.role == 'distributor' and user.area:
            # For distributors, get products from dealer_wise_stock_details (their own stock)
            products = db_service.get_products_from_dealer_stock(user.area)
        else:
            # Fallback: use Product table only if no area or not MR/distributor
            warehouse = db_service.get_warehouse_by_area(user.area) if user.area else None
            if warehouse:
                products = db_service.get_products_by_warehouse(warehouse.id)
            else:
                # Get all products if no warehouse
                from app.models import Product
                products = Product.query.all()
        
        if not products:
            response = "No products are currently available in your area."
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Place Order', 'action': 'place_order'},
                    {'text': 'Track Order', 'action': 'track_order'},
                    {'text': 'Get Help', 'action': 'help'}
                ]
            }), 200
        else:
            # Build simple response message
            response = f"Here are all the products available in your area:\n\n"
            response += f"**Total Products:** {len(products)}\n\n"
            response += "Would you like to place an order for any of these products?"
            
            # Build product list for table display
            product_list = []
            for product in products:
                # Handle both dictionary (from dealer stock) and Product objects
                if isinstance(product, dict):
                    # Dictionary from dealer stock
                    price = product.get('sales_price', product.get('price', 0))
                    available = product.get('available_quantity', 0)
                    product_name = product.get('product_name', '')
                    product_code = product.get('product_code', '')
                else:
                    # Product object
                    price = product.sales_price if hasattr(product, 'sales_price') and product.sales_price else product.price
                    available = 0  # Will be in dealer stock
                    product_name = product.product_name
                    product_code = str(product.id)
                
                product_list.append({
                    'product_name': product_name,
                    'product_code': product_code,
                    'sales_price': float(price) if price else 0.0,
                    'available_for_sale': int(available) if available else 0
                })
            
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'products': product_list,
                'show_product_table': True,  # Flag to show table in frontend
                'action_buttons': [
                    {'text': 'Place Order', 'action': 'place_order'},
                    {'text': 'Track Order', 'action': 'track_order'},
                    {'text': 'Get Help', 'action': 'help'}
                ]
            }), 200
        
    except Exception as e:
        logger.error(f"Error handling product info/query: {str(e)}")
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error retrieving product information. Please try again.'}), user), 500

def handle_distributor_analytics(user_message, user, db_service, llm_service):
    """Handle dynamic database queries and analytics for distributors"""
    try:
        from app.models import Order, OrderItem, Product, User
        
        # Build warehouse context using user's area
        area = user.area if user.area else 'N/A'
        
        # Get relevant data for analytics - filter orders by user's area
        # Explicitly specify join condition to avoid ambiguous foreign key error
        warehouse_orders = Order.query.join(User, Order.mr_id == User.id).filter(User.area == area).all()
        
        # Create context for LLM
        analytics_context = f"""
Area: {area}
Total Orders: {len(warehouse_orders)}
Recent Orders: {min(10, len(warehouse_orders))} orders
Order Statuses: {', '.join(set(o.status for o in warehouse_orders[:10]))}

Available Tables and Data:
- orders: Order information (order_id, status, total_amount, order_date, user_email, user_id)
- order_items: Order line items (product_code, quantity, unit_price, total_price)
- products: Product information (product_code, product_name, price_of_product, available_for_sale, warehouse_id)
- pending_order_products: Pending orders waiting for stock (product_code, requested_quantity, customer, status)
- users: Customer information (name, email, user_type, area)

User Query: {user_message}
"""
        
        # Use LLM to generate SQL or provide insights
        llm_prompt = f"""You are an AI assistant helping a distributor query their database.
        
Context:
{analytics_context}

Based on the user's query, provide a helpful response using the database information available.
You can answer questions about:
- Order statistics and totals
- Product availability and inventory
- Customer information
- Revenue and sales analytics
- Pending orders

Respond in a conversational but informative way, providing specific numbers and insights.
If the query requires data that might not be available, provide a helpful alternative response.

User Query: "{user_message}"

Provide a direct answer:"""
        
        response_obj = llm_service.client.chat.completions.create(
            model='llama-3.3-70b-versatile',
            messages=[{"role": "user", "content": llm_prompt}],
            temperature=0.2,
            max_tokens=500
        )
        
        response_text = response_obj.choices[0].message.content.strip()
        
        # Try to enhance with actual data
        if 'how many orders' in user_message.lower() or 'count of orders' in user_message.lower():
            response_text += f"\n\n**Actual Data:**\n"
            response_text += f"• Total orders in warehouse: {len(warehouse_orders)}\n"
            
            # Count by status
            from collections import Counter
            status_counts = Counter(o.status for o in warehouse_orders)
            for status, count in status_counts.items():
                response_text += f"• {status}: {count}\n"
        
        save_conversation(user.id, user_message, response_text)
        return jsonify({
            'response': response_text,
            'action_buttons': [
                {'text': 'View Orders', 'action': 'track_order'}
            ]
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling distributor analytics: {str(e)}")
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error processing your analytics request. Please try again.'}), user), 500

def handle_web_search(user_message, user, context_data):
    """Handle web search requests"""
    try:
        web_search_service = get_web_search_service()
        search_result = web_search_service.search(user_message)
        
        if search_result['success']:
            response = f"Here's what I found about '{user_message}':\n\n"
            response += search_result['content']
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'sources': search_result.get('sources', [])
            }), 200
        else:
            response = f"I couldn't find specific information about '{user_message}'. Please try rephrasing your question or ask about our products and services."
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Get Help', 'action': 'help'}
                ]
            }), 200
            
    except Exception as e:
        logger.error(f"Error handling web search: {str(e)}")
        return ensure_action_buttons(jsonify({'response': 'Sorry, I encountered an error with the web search.'}), user), 500

def handle_stock_confirmation(user_message, user, context_data):
    """Handle stock confirmation requests for dealers"""
    try:
        import re
        from app.models import DealerWiseStockDetails
        
        db_service = get_db_service()
        stock_service = get_stock_management_service()
        
        message_lower = user_message.lower()
        
        # Check if user wants to see pending stock
        if any(keyword in message_lower for keyword in ['show', 'list', 'display', 'pending', 'view']):
            # Check if user specified an invoice_id
            invoice_id = None
            invoice_match = re.search(r'invoice[_\s]*id[:\s]*([A-Za-z0-9_-]+)', message_lower, re.IGNORECASE)
            if invoice_match:
                invoice_id = invoice_match.group(1).strip()
            
            # Check if user specified a date filter
            date_filter = None
            date_match = re.search(r'date[:\s]*(\d{4}-\d{2}-\d{2})', message_lower, re.IGNORECASE)
            if date_match:
                date_filter = date_match.group(1).strip()
            
            # Get pending stock arrivals (with optional filters)
            result = stock_service.get_pending_stock_arrivals(user.unique_id, invoice_id=invoice_id, date_filter=date_filter)
            
            if not result['success'] or not result['stocks']:
                if invoice_id:
                    response = f"You have no pending stock arrivals for invoice ID: {invoice_id}."
                elif date_filter:
                    response = f"You have no pending stock arrivals for date: {date_filter}."
                else:
                    response = "You have no pending stock arrivals to confirm."
                save_conversation(user.id, user_message, response)
                return jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'View Orders', 'action': 'track_order'}
                    ]
                }), 200
            
            # Create a simplified response message
            response = f"**You have {result['count']} pending stock arrival(s) to confirm:**\n\n"
            response += "**Use the form below to view and confirm stock:**\n"
            
            # Remove purchase_price from stocks data before sending to frontend
            stocks_for_display = []
            for stock in result['stocks']:
                stock_display = {k: v for k, v in stock.items() if k != 'purchase_price'}
                stocks_for_display.append(stock_display)
            
            save_conversation(user.id, user_message, response)
            return jsonify({
                'response': response,
                'stocks': stocks_for_display,
                'invoice_ids': result.get('invoice_ids', []),
                'invoice_dates': result.get('dispatch_dates', []),  # Renamed to invoice_dates
                'dispatch_dates': result.get('dispatch_dates', []),  # Keep for backward compatibility
                'interactive_stock_confirmation': True,  # Flag for frontend to render interactive UI
                'show_stock_table': True,  # Flag to show table in form
                'action_buttons': [
                    {'text': 'View Orders', 'action': 'track_order'}
                ]
            }), 200
        
        # Check if user wants to confirm stock
        # Pattern: "confirm stock <id>" or "confirm stock <id> received <quantity>"
        confirm_pattern = r'confirm\s+stock\s+(\d+)'
        match = re.search(confirm_pattern, message_lower)
        
        if match:
            stock_id = int(match.group(1))
            
            # Try to extract received quantity
            received_qty_match = re.search(r'received\s+(\d+)', message_lower)
            received_quantity = int(received_qty_match.group(1)) if received_qty_match else None
            
            # Try to extract adjustment reason - improved regex to capture everything after "reason"
            # Pattern: "reason" followed by whitespace, then capture everything until end of string
            reason_match = re.search(r'reason\s+(.+)', message_lower, re.IGNORECASE | re.DOTALL)
            adjustment_reason = reason_match.group(1).strip() if reason_match else None
            
            # Log extracted values for debugging
            logger.info(f"Stock confirmation - ID: {stock_id}, Received: {received_quantity}, Reason: {adjustment_reason}")
            
            # Confirm stock arrival
            result = stock_service.confirm_stock_arrival(
                stock_id,
                user.id,
                received_quantity,
                adjustment_reason
            )
            
            if result['success']:
                stock_detail = result['stock_detail']
                response = f"✅ **Stock confirmed successfully!**\n\n"
                response += f"**Product:** {stock_detail['product_name']}\n"
                response += f"**Quantity Received:** {stock_detail.get('received_quantity', stock_detail['quantity'])} units\n"
                
                if result.get('quantity_adjusted'):
                    response += f"⚠️ **Quantity Adjusted:** Yes\n"
                    response += f"**Reason:** {stock_detail.get('adjustment_reason', 'N/A')}\n"
                    response += f"**Note:** An email has been sent to the company about this adjustment.\n"
                
                response += f"\n**Status:** {stock_detail['status']}\n"
                response += f"**Confirmed At:** {stock_detail.get('confirmed_at', 'N/A')}\n"
                
                # Check if there are more pending stocks
                pending_result = stock_service.get_pending_stock_arrivals(user.unique_id)
                
                if pending_result['success'] and pending_result['count'] > 0:
                    response += f"\n**You have {pending_result['count']} more pending stock arrival(s) to confirm.**\n"
                
                save_conversation(user.id, user_message, response)
                
                response_data = {
                    'response': response,
                    'stock_detail': stock_detail,
                    'action_buttons': [
                        {'text': 'Show Pending Stock', 'action': 'show_pending_stock'},
                        {'text': 'View Orders', 'action': 'track_order'}
                    ]
                }
                
                # Show interactive form for remaining stocks if any
                if pending_result['success'] and pending_result['count'] > 0:
                    response_data['interactive_stock_confirmation'] = True
                    response_data['stocks'] = pending_result['stocks']
                    response_data['invoice_ids'] = pending_result.get('invoice_ids', [])
                    response_data['dispatch_dates'] = pending_result.get('dispatch_dates', [])
                    response_data['show_stock_table'] = True
                
                return jsonify(response_data), 200
            else:
                response = f"❌ **Error:** {result['message']}"
                save_conversation(user.id, user_message, response)
                return jsonify({
                    'response': response,
                    'action_buttons': [
                        {'text': 'Show Pending Stock', 'action': 'show_pending_stock'},
                        {'text': 'Get Help', 'action': 'help'}
                    ]
                }), 200
        
        # If no pattern matched, show help
        response = "**Stock Confirmation Help:**\n\n"
        response += "To view pending stock arrivals:\n"
        response += "• Type: `show pending stock` or `list pending arrivals`\n\n"
        response += "To confirm stock:\n"
        response += "• Type: `confirm stock <stock_id>` (e.g., 'confirm stock 1')\n"
        response += "• If quantity differs: `confirm stock <stock_id> received <quantity>` (e.g., 'confirm stock 1 received 95')\n"
        response += "• With reason: `confirm stock <stock_id> received <quantity> reason <reason>`\n\n"
        response += "**Example:**\n"
        response += "• 'show pending stock' - View all pending arrivals\n"
        response += "• 'confirm stock 1' - Confirm stock ID 1 with sent quantity\n"
        response += "• 'confirm stock 1 received 95' - Confirm stock ID 1 with 95 units received\n"
        response += "• 'confirm stock 1 received 95 reason damaged during transit' - Confirm with quantity and reason\n"
        
        save_conversation(user.id, user_message, response)
        return jsonify({
            'response': response,
            'action_buttons': [
                {'text': 'Show Pending Stock', 'action': 'show_pending_stock'},
                {'text': 'View Orders', 'action': 'track_order'}
            ]
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling stock confirmation: {str(e)}")
        import traceback
        traceback.print_exc()
        response = f"Sorry, I encountered an error processing your stock confirmation request: {str(e)}"
        save_conversation(user.id, user_message, response)
        return jsonify({
            'response': response,
            'action_buttons': [
                {'text': 'Show Pending Stock', 'action': 'show_pending_stock'},
                {'text': 'Get Help', 'action': 'help'}
            ]
        }), 500

def handle_general_conversation(user_message, user, context_data):
    """Handle general conversation using LLM"""
    try:
        llm_service = get_llm_service()
        
        if not llm_service.client:
            response = "I'm here to help you with orders, tracking, and company information. How can I assist you today?"
            save_conversation(user.id, user_message, response)
            
            # Build action buttons
            action_buttons = [
                {'text': 'Place Order', 'action': 'place_order'},
                {'text': 'Track Order', 'action': 'track_order'},
                {'text': 'Company Info', 'action': 'company_info'}
            ]
            
            # For distributors (dealers), check for pending stocks and add button if needed
            if user.role == 'distributor':
                stock_service = get_stock_management_service()
                pending_result = stock_service.get_pending_stock_arrivals(user.unique_id)
                
                if pending_result['success'] and pending_result['count'] > 0:
                    # Add "Pending Stocks" button at the beginning
                    action_buttons.insert(0, {'text': 'Pending Stocks', 'action': 'pending_stocks'})
            
            return jsonify({
                'response': response,
                'action_buttons': action_buttons
            }), 200
        
        # Generate contextual response
        context_prompt = f"""You are Quantum Blue's AI assistant for HV (Powered by Quantum Blue AI). 
        
User: {user_message}

User Context:
- Name: {user.name}
- Type: {user.role}
- Role: {user.role or 'N/A'}
- Area: {user.area or 'N/A'}

Your task:
1. Provide a helpful, friendly response
2. Guide the user toward our main services (ordering, tracking, company info)
3. Be conversational but professional
4. If the user seems confused, offer to help with specific tasks

Respond naturally and helpfully."""

        response_obj = llm_service.client.chat.completions.create(
            model=current_app.config.get('GROQ_MODEL', 'mixtral-8x7b-32768'),
            messages=[{"role": "user", "content": context_prompt}],
            temperature=0.7,
            max_tokens=500
        )
        
        response = response_obj.choices[0].message.content
        save_conversation(user.id, user_message, response)
        
        # Build action buttons
        action_buttons = [
            {'text': 'Place Order', 'action': 'place_order'},
            {'text': 'Track Order', 'action': 'track_order'},
            {'text': 'Company Info', 'action': 'company_info'}
        ]
        
        # For distributors (dealers), check for pending stocks and add button if needed
        if user.role == 'distributor':
            stock_service = get_stock_management_service()
            pending_result = stock_service.get_pending_stock_arrivals(user.unique_id)
            
            if pending_result['success'] and pending_result['count'] > 0:
                # Add "Pending Stocks" button at the beginning
                action_buttons.insert(0, {'text': 'Pending Stocks', 'action': 'pending_stocks'})
        
        action_buttons.append({'text': 'Get Help', 'action': 'help'})
        
        return jsonify({
            'response': response,
            'action_buttons': action_buttons
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling general conversation: {str(e)}")
        response = "I'm here to help you with orders, tracking, and company information. How can I assist you today?"
        save_conversation(user.id, user_message, response)
        
        # Build action buttons
        action_buttons = [
            {'text': 'Place Order', 'action': 'place_order'},
            {'text': 'Track Order', 'action': 'track_order'},
            {'text': 'Company Info', 'action': 'company_info'}
        ]
        
        # For distributors (dealers), check for pending stocks and add button if needed
        if user.role == 'distributor':
            stock_service = get_stock_management_service()
            pending_result = stock_service.get_pending_stock_arrivals(user.unique_id)
            
            if pending_result['success'] and pending_result['count'] > 0:
                # Add "Pending Stocks" button at the beginning
                action_buttons.insert(0, {'text': 'Pending Stocks', 'action': 'pending_stocks'})
        
        return jsonify({
            'response': response,
            'action_buttons': action_buttons
        }), 200

def generate_welcome_message(user):
    """Generate personalized welcome message based on user type"""
    if user.role == 'company':
        return f"""Welcome {user.name}! 👋

**HV Company Analytics System**

I'm your database reporting assistant. I can help you:

• **Generate Database Reports** - Export any table to CSV
• **Filter Columns** - Select specific data you need
• **Email Delivery** - Receive reports directly in your inbox

**Available Reports:**
- Users, Orders, Order Items, Products
- Customers, Cart Items, FOC Schemes
- Dealer Stock, Pending Orders, Email Logs

Type **"generate report"** to get started or **"help"** for more information."""
    elif user.role == 'customer':
        return f"Welcome back, {user.name}! I'm here to help you with your orders and answer any questions about our products. What would you like to do today?"
    elif user.role == 'mr':
        return f"Hello {user.name}! As a Medical Representative, you can place orders for your clients and track deliveries. How can I assist you today?"
    elif user.role == 'distributor':
        # Check for pending stocks to confirm
        stock_service = get_stock_management_service()
        pending_result = stock_service.get_pending_stock_arrivals(user.unique_id)
        
        welcome_msg = f"Welcome {user.name}! As a Distributor, you can manage orders, confirm deliveries, and track inventory."
        
        if pending_result['success'] and pending_result['count'] > 0:
            welcome_msg += f"\n\n📦 **Important:** You have {pending_result['count']} pending stock arrival(s) to confirm. Type 'show pending stock' or 'show pending orders' to view and confirm them."
        
        welcome_msg += "\n\nWhat would you like to do?"
        return welcome_msg
    elif user.role == 'pharmacy':
        return f"Hello {user.name}! As a Pharmacy, you can place orders and track your deliveries. How can I help you today?"
    else:
        return f"Welcome back, {user.name}! I'm here to help you with your orders and answer any questions. What would you like to do today?"

def save_conversation(user_id, user_message, bot_response):
    """
    Save conversation to database.
    IMPORTANT: Always saves in English (original language).
    Translation happens only when sending to user, not when saving to database.
    """
    try:
        db_service = get_db_service()
        session_id = session.get('session_id')
        
        if session_id:
            # Always save original English text to database
            # Translation is only for display to user, not for storage
            db_service.save_conversation(
                user_id=user_id,
                session_id=session_id,
                user_message=user_message,
                bot_response=bot_response  # This should be the original English response
            )
    except Exception as e:
        logger.error(f"Error saving conversation: {str(e)}")

def handle_select_customer(user, context_data):
    """Handle select customer action - show customer selection form"""
    try:
        if user.role != 'mr':
            return jsonify({
                'response': 'Only Medical Representatives can select customers.',
                'action_buttons': []
            }), 200
        
        # Get customers for this MR
        customers = Customer.query.filter_by(
            mr_unique_id=user.unique_id,
            is_active=True
        ).all()
        
        if not customers:
            response = "No customers are assigned to you. Please add a new customer or contact support to assign customers."
            save_conversation(user.id, "select customer", response)
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Add New Customer', 'action': 'add_new_customer'},
                    {'text': 'Get Help', 'action': 'help'}
                ]
            }), 200
        
        # Build customer list (only name and ID)
        customer_list = []
        for customer in customers:
            customer_list.append({
                'id': customer.id,
                'unique_id': customer.unique_id,
                'name': customer.name
            })
        
        response = "Please select the customer for whom you want to place an order:"
        
        save_conversation(user.id, "select customer", response)
        
        return jsonify({
            'response': response,
            'customers': customer_list,
            'interactive_customer_selection': True,
            'show_customer_table': True,  # Flag to show table outside selection box
            'action_buttons': []  # No action buttons - removed Add New Customer and Cancel
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling select customer: {str(e)}")
        return jsonify({
            'response': 'Sorry, I encountered an error. Please try again.',
            'action_buttons': []
        }), 200

def handle_add_new_customer(user, context_data):
    """Handle add new customer action - show add customer form"""
    try:
        if user.role != 'mr':
            return jsonify({
                'response': 'Only Medical Representatives can add customers.',
                'action_buttons': []
            }), 200
        
        response = "Please fill in the details below to add a new customer."
        
        save_conversation(user.id, "add new customer", response)
        
        return jsonify({
            'response': response,
            'interactive_add_customer': True
        }), 200
        
    except Exception as e:
        logger.error(f"Error handling add new customer: {str(e)}")
        return jsonify({
            'response': 'Sorry, I encountered an error. Please try again.',
            'action_buttons': []
        }), 200

@chatbot_bp.route('/select_customer', methods=['POST'])
def select_customer():
    """Select customer for MR order"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'mr':
            return jsonify({'error': 'Only MRs can select customers'}), 403
        
        data = request.get_json()
        customer_id = data.get('customer_id')
        
        if not customer_id:
            return jsonify({'error': 'Customer ID is required'}), 400
        
        # Verify customer belongs to this MR
        customer = Customer.query.filter_by(
            id=customer_id,
            mr_unique_id=user.unique_id,
            is_active=True
        ).first()
        
        if not customer:
            return jsonify({'error': 'Customer not found or not assigned to you'}), 404
        
        # Store selected customer in session
        session['selected_customer_id'] = customer.id
        session['selected_customer_unique_id'] = customer.unique_id
        
        # Get available products for the MR
        db_service = get_db_service()
        if user.area:
            products = db_service.get_products_from_dealer_stock(user.area)
        else:
            from app.models import Product
            products = Product.query.all()
        
        # Build product list for interactive UI with FOC information using helper
        pricing_service = get_pricing_service()
        product_list = build_product_list_with_foc(products, pricing_service)
        
        # Create response message with customer info
        response = f"• Great! I can help you place an order.\n• **Ordering for:** {customer.name} ({customer.unique_id})\n• Please use the product selection form below to select products and quantities."
        
        # Save conversation
        save_conversation(user.id, f"Selected customer: {customer.name}", response)
        
        return jsonify({
            'success': True,
            'message': f'Customer {customer.name} selected successfully',
            'customer': {
                'id': customer.id,
                'unique_id': customer.unique_id,
                'name': customer.name,
                'email': customer.email,
                'phone': customer.phone
            },
            'response': response,
            'products': product_list,
            'interactive_product_selection': True,  # Flag for frontend to render interactive UI
            'show_product_table': True,  # Flag to show table outside selection box
            'action_buttons': [{'text': 'Change Customer', 'action': 'change_customer'}]
        }), 200
        
    except Exception as e:
        logger.error(f"Error selecting customer: {str(e)}")
        return jsonify({'error': 'Error selecting customer'}), 500

@chatbot_bp.route('/select_order', methods=['POST'])
def select_order():
    """Select order for distributor or MR to view details"""
    from app.input_validation import validate_and_sanitize_order_id
    try:
        user_id = session.get('user_id')
        if not user_id:
            logger.warning("select_order: User not logged in")
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user:
            logger.warning(f"select_order: User {user_id} not found")
            return jsonify({'error': 'User not found'}), 404
        
        # Allow both distributors and MRs to select orders
        if user.role not in ['distributor', 'mr']:
            logger.warning(f"select_order: User {user_id} has invalid role: {user.role}")
            return jsonify({'error': 'Only distributors and MRs can use this feature'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        order_id_raw = data.get('order_id')
        if not order_id_raw:
            logger.warning(f"select_order: Order ID not provided by user {user_id}")
            return jsonify({'error': 'Order ID is required'}), 400
        
        order_id = validate_and_sanitize_order_id(order_id_raw)
        if not order_id:
            return jsonify({'error': 'Invalid order ID format'}), 400
        
        logger.info(f"select_order: User {user_id} ({user.role}) requesting order {order_id}")
        
        # Get order details
        enhanced_order_service = get_enhanced_order_service()
        order = Order.query.filter_by(order_id=order_id).first()
        
        if not order:
            logger.warning(f"select_order: Order {order_id} not found")
            return jsonify({'error': 'Order not found'}), 404
        
        # Verify order access based on user role
        if user.role == 'distributor':
            # Distributors can only see orders in their area
            if order.mr and order.mr.area != user.area:
                logger.warning(f"select_order: Order {order_id} not in distributor {user_id}'s area. Order area: {order.mr.area}, User area: {user.area}")
                return jsonify({'error': 'This order is not in your area'}), 403
        elif user.role == 'mr':
            # MRs can only see their own orders
            if order.mr_id != user.id:
                logger.warning(f"select_order: Order {order_id} does not belong to MR {user_id}. Order MR: {order.mr_id}, User: {user.id}")
                return jsonify({'error': 'You can only view your own orders'}), 403
        
        logger.info(f"select_order: Access granted for user {user_id} to order {order_id}")
        
        # Get order items with FOC and lot numbers
        items_list = []
        total_items = 0
        
        # Get dealers in MR's area for lot number lookup
        dealers_in_area = []
        if order.mr and order.mr.area:
            dealers_in_area = User.query.filter_by(
                role='distributor',
                area=order.mr.area
            ).all()
        dealer_unique_ids = [d.unique_id for d in dealers_in_area] if dealers_in_area else []
        
        for item in order.order_items:
            # Use adjusted_quantity if available (after dealer confirmation), otherwise use original quantity
            # This ensures we display the actual dispatched quantity, not the original ordered quantity
            actual_quantity = item.adjusted_quantity if item.adjusted_quantity is not None else item.quantity
            free_qty = item.free_quantity or 0
            # When quantity is adjusted, FOC may have been moved to pending, so use current free_quantity
            total_qty = actual_quantity + free_qty
            total_items += total_qty
            
            # Get lot number - prefer adjusted_lot_number if available, otherwise from DealerWiseStockDetails
            lot_number = None
            expiry_date = None
            
            # First check if adjusted_lot_number exists (from dealer confirmation)
            if item.adjusted_lot_number:
                lot_number = item.adjusted_lot_number
            
            # If no adjusted lot number, try to get from DealerWiseStockDetails
            if not lot_number and dealer_unique_ids:
                # Get the first available stock detail with lot number for this product
                # Use FEFO (First Expiry First Out) - earliest expiry first
                from sqlalchemy import case
                stock_detail = DealerWiseStockDetails.query.filter(
                    DealerWiseStockDetails.product_code == item.product_code,
                    DealerWiseStockDetails.status == 'confirmed',
                    DealerWiseStockDetails.dealer_unique_id.in_(dealer_unique_ids),
                    DealerWiseStockDetails.blocked_quantity > 0
                ).order_by(
                    case(
                        (DealerWiseStockDetails.expiry_date.is_(None), 1),
                        else_=0
                    ),
                    DealerWiseStockDetails.expiry_date.asc()
                ).first()
                
                if stock_detail:
                    lot_number = stock_detail.lot_number
                    expiry_date = stock_detail.expiry_date.strftime('%Y-%m-%d') if stock_detail.expiry_date else None
            
            # Prefer adjusted_expiry_date if available
            if item.adjusted_expiry_date:
                expiry_date = item.adjusted_expiry_date.strftime('%Y-%m-%d')
            
            # Recalculate total_price based on adjusted quantity if quantity was adjusted
            unit_price = float(item.unit_price) if item.unit_price else 0.0
            if item.adjusted_quantity is not None and item.adjusted_quantity != item.quantity:
                # Quantity was adjusted, recalculate total_price based on adjusted quantity
                total_price = float(actual_quantity * unit_price)
            else:
                # Use original total_price if quantity wasn't adjusted
                total_price = float(item.total_price) if item.total_price else 0.0
            
            items_list.append({
                'id': item.id,  # Include item ID for editing
                'product_name': item.product.product_name,
                'product_code': item.product_code,
                'quantity': actual_quantity,  # Use adjusted quantity if available
                'original_quantity': item.quantity,  # Store original quantity for comparison
                'free_quantity': free_qty,
                'total_quantity': total_qty,
                'unit_price': unit_price,
                'total_price': total_price,  # Recalculated if quantity was adjusted
                'lot_number': lot_number,  # Prefer adjusted_lot_number, fallback to stock details
                'expiry_date': expiry_date,  # Prefer adjusted_expiry_date, fallback to stock details
                'adjusted_quantity': item.adjusted_quantity,  # Include for reference
                'adjustment_reason': item.adjustment_reason  # Include adjustment reason if available
            })
        
        # Get customer info if available
        customer_name = None
        customer_id = None
        if order.customer_id:
            # Try to get customer from relationship first
            if order.customer:
                customer_name = f"{order.customer.name} ({order.customer.unique_id})"
                customer_id = order.customer.unique_id
            else:
                # If relationship not loaded, query directly
                from app.models import Customer
                customer = Customer.query.get(order.customer_id)
                if customer:
                    customer_name = f"{customer.name} ({customer.unique_id})"
                    customer_id = customer.unique_id
        elif order.customer_unique_id:
            # Fallback: try to get customer by unique_id
            from app.models import Customer
            customer = Customer.query.filter_by(unique_id=order.customer_unique_id).first()
            if customer:
                customer_name = f"{customer.name} ({customer.unique_id})"
                customer_id = customer.unique_id
        
        # Find any pending orders that were fulfilled by this order
        pending_sources = PendingOrderProducts.query.filter_by(fulfilled_order_id=order.order_id).all()
        pending_source_orders = []
        for p in pending_sources:
            original_order = None
            if p.original_order_id:
                original_order = Order.query.filter_by(order_id=p.original_order_id).first()
            original_date = original_order.created_at.strftime('%Y-%m-%d %H:%M:%S') if (original_order and original_order.created_at) else None
            pending_source_orders.append({
                'pending_id': p.id,
                'original_order_id': p.original_order_id,
                'product_code': p.product_code,
                'product_name': p.product_name,
                'requested_quantity': p.requested_quantity,
                'original_order_date': original_date
            })
        
        # Find any pending orders that were created FROM this order (when order was partially dispatched)
        pending_orders_from_this = PendingOrderProducts.query.filter_by(original_order_id=order.order_id).all()
        pending_orders_created = []
        for p in pending_orders_from_this:
            fulfilled_order = None
            fulfilled_order_date = None
            if p.fulfilled_order_id:
                fulfilled_order = Order.query.filter_by(order_id=p.fulfilled_order_id).first()
                fulfilled_order_date = fulfilled_order.created_at.strftime('%Y-%m-%d %H:%M:%S') if (fulfilled_order and fulfilled_order.created_at) else None
            pending_orders_created.append({
                'pending_id': p.id,
                'product_code': p.product_code,
                'product_name': p.product_name,
                'requested_quantity': p.requested_quantity,
                'original_foc_quantity': p.original_foc_quantity or 0,
                'total_pending_quantity': (p.requested_quantity or 0) + (p.original_foc_quantity or 0),
                'status': p.status,
                'fulfilled_order_id': p.fulfilled_order_id,
                'fulfilled_order_date': fulfilled_order_date,
                'created_at': p.created_at.strftime('%Y-%m-%d %H:%M:%S') if p.created_at else None
            })
        
        # Get delivery partner info if available
        delivery_partner_name = None
        delivery_partner_email = None
        delivery_partner_phone = None
        delivery_partner_unique_id = None
        if order.delivery_partner_id:
            delivery_partner = User.query.get(order.delivery_partner_id)
            if delivery_partner:
                delivery_partner_name = delivery_partner.name
                delivery_partner_email = delivery_partner.email
                delivery_partner_phone = delivery_partner.phone
                delivery_partner_unique_id = delivery_partner.unique_id
        elif order.delivery_partner_unique_id:
            delivery_partner = User.query.filter_by(unique_id=order.delivery_partner_unique_id).first()
            if delivery_partner:
                delivery_partner_name = delivery_partner.name
                delivery_partner_email = delivery_partner.email
                delivery_partner_phone = delivery_partner.phone
                delivery_partner_unique_id = delivery_partner.unique_id
        
        # Recalculate order totals based on adjusted quantities if any items were adjusted
        # Sum up all item total prices (which are already recalculated based on adjusted quantities)
        recalculated_subtotal = sum(item.get('total_price', 0) for item in items_list)
        # Check if any item has adjusted_quantity that differs from original_quantity
        has_adjustments = any(item.get('adjusted_quantity') is not None and 
                             item.get('adjusted_quantity') != item.get('original_quantity', 0) 
                             for item in items_list)
        
        # Use recalculated values if quantities were adjusted, otherwise use original order values
        if has_adjustments:
            tax_rate = float(order.tax_rate) if hasattr(order, 'tax_rate') and order.tax_rate else 0.05
            recalculated_tax = recalculated_subtotal * tax_rate
            recalculated_total = recalculated_subtotal + recalculated_tax
            subtotal = recalculated_subtotal
            tax_amount = recalculated_tax
            total_amount = recalculated_total
        else:
            subtotal = float(order.subtotal) if hasattr(order, 'subtotal') and order.subtotal else 0.0
            tax_rate = float(order.tax_rate) if hasattr(order, 'tax_rate') and order.tax_rate else 0.05
            tax_amount = float(order.tax_amount) if hasattr(order, 'tax_amount') and order.tax_amount else 0.0
            total_amount = float(order.total_amount) if order.total_amount else 0.0
        
        # Build detailed response including a timeline-friendly set of timestamps
        order_details = {
            'order_id': order.order_id,
            'mr_name': order.mr.name if order.mr else 'N/A',
            'mr_email': order.mr.email if order.mr else 'N/A',
            'mr_phone': order.mr.phone if order.mr else 'N/A',
            'status': order.status,
            'status_display': (order.status or 'Unknown').replace('_', ' ').title(),
            'order_stage': getattr(order, 'order_stage', None),
            'subtotal': subtotal,
            'tax_rate': tax_rate,
            'tax_amount': tax_amount,
            'total_amount': total_amount,
            'total_items': total_items,
            # Base timestamps
            'order_date': order.created_at.strftime('%Y-%m-%d') if order.created_at else 'N/A',
            'order_time': order.created_at.strftime('%H:%M') if order.created_at else 'N/A',
            'order_datetime': order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else 'N/A',
            # Timeline specific fields (ISO strings so frontend can render nicely)
            'placed_at': order.created_at.isoformat() if order.created_at else None,
            'distributor_confirmed_at': order.distributor_confirmed_at.isoformat() if order.distributor_confirmed_at else None,
            'delivered_at': order.delivered_at.isoformat() if order.delivered_at else None,
            'last_updated_at': order.updated_at.isoformat() if order.updated_at else None,
            # Pending-order linkage info
            'is_fulfilled_pending_order': True if pending_sources else False,
            'pending_source_orders': pending_source_orders,
            # Pending orders created FROM this order (when partially dispatched)
            'has_pending_orders': True if pending_orders_created else False,
            'pending_orders_created': pending_orders_created,
            # Delivery partner info
            'delivery_partner_name': delivery_partner_name,
            'delivery_partner_email': delivery_partner_email,
            'delivery_partner_phone': delivery_partner_phone,
            'delivery_partner_unique_id': delivery_partner_unique_id,
            'area': order.mr.area if order.mr else 'N/A',
            'customer_name': customer_name,
            'customer_id': customer_id,
            'items': items_list,
            'can_confirm': order.status == 'pending' and user.role == 'distributor',
            'can_reject': order.status == 'pending' and user.role == 'distributor',
            'can_cancel': order.status in ['pending', 'draft'] and user.role == 'mr'
        }
        
        return jsonify({
            'success': True,
            'order': order_details
        }), 200
        
    except Exception as e:
        logger.error(f"Error selecting order: {str(e)}")
        return jsonify({'error': 'Error selecting order'}), 500

@chatbot_bp.route('/confirm_order_action', methods=['POST'])
def confirm_order_action():
    """Confirm order by distributor"""
    try:
        from app.input_validation import (
            validate_and_sanitize_order_id, validate_quantity, 
            sanitize_string, MAX_LENGTHS, sanitize_dict
        )
        
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Only distributors can confirm orders'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        # Validate and sanitize order_id
        order_id_raw = data.get('order_id')
        if not order_id_raw:
            return jsonify({'error': 'Order ID is required'}), 400
        
        order_id = validate_and_sanitize_order_id(order_id_raw)
        if not order_id:
            return jsonify({'error': 'Invalid order ID format'}), 400
        
        # Validate and sanitize delivery_partner_id if provided
        delivery_partner_id = None
        delivery_partner_id_raw = data.get('delivery_partner_id')
        if delivery_partner_id_raw is not None:
            try:
                delivery_partner_id = int(delivery_partner_id_raw)
                if delivery_partner_id <= 0:
                    delivery_partner_id = None
                else:
                    # Verify delivery partner exists and is valid
                    delivery_partner = User.query.get(delivery_partner_id)
                    if not delivery_partner or delivery_partner.role != 'delivery_partner':
                        return jsonify({'error': 'Invalid delivery partner selected'}), 400
            except (ValueError, TypeError):
                delivery_partner_id = None
        
        # Validate and sanitize item_edits
        item_edits_raw = data.get('item_edits')  # {item_id: {'quantity': int, 'expiry_date': str, 'reason': str}}
        
        # Handle null/None item_edits - convert to empty dict so backend uses original quantities
        if item_edits_raw is None:
            item_edits = {}
        elif not isinstance(item_edits_raw, dict):
            item_edits = {}
        else:
            # Sanitize item_edits dictionary
            item_edits = {}
            for item_id_str, edits in item_edits_raw.items():
                if not isinstance(edits, dict):
                    continue
                
                try:
                    item_id = int(item_id_str)
                    if item_id <= 0:
                        continue
                    
                    sanitized_edit = {}
                    
                    # Validate quantity
                    if 'quantity' in edits:
                        qty = edits['quantity']
                        if validate_quantity(qty):
                            sanitized_edit['quantity'] = int(qty)
                    
                    # Sanitize lot_number
                    if 'lot_number' in edits and edits['lot_number']:
                        lot = sanitize_string(str(edits['lot_number']), max_length=MAX_LENGTHS['lot_number'])
                        if lot:
                            sanitized_edit['lot_number'] = lot
                    
                    # Validate expiry_date (format: YYYY-MM-DD)
                    if 'expiry_date' in edits and edits['expiry_date']:
                        expiry = sanitize_string(str(edits['expiry_date']), max_length=10)
                        if expiry and re.match(r'^\d{4}-\d{2}-\d{2}$', expiry):
                            sanitized_edit['expiry_date'] = expiry
                    
                    # Sanitize reason
                    if 'reason' in edits and edits['reason']:
                        reason = sanitize_string(str(edits['reason']), max_length=MAX_LENGTHS['reason'])
                        if reason:
                            sanitized_edit['reason'] = reason
                    
                    if sanitized_edit:
                        item_edits[item_id] = sanitized_edit
                        
                except (ValueError, TypeError):
                    continue
        
        # Convert expiry_date strings to date objects and item_id to int if provided
        if item_edits:
            from datetime import datetime
            converted_edits = {}
            for item_id_str, edits in item_edits.items():
                try:
                    item_id = int(item_id_str)
                    converted_edit = edits.copy()
                    if 'expiry_date' in converted_edit and converted_edit['expiry_date']:
                        try:
                            converted_edit['expiry_date'] = datetime.strptime(converted_edit['expiry_date'], '%Y-%m-%d').date()
                        except:
                            pass
                    converted_edits[item_id] = converted_edit
                except (ValueError, TypeError):
                    pass
            item_edits = converted_edits
        
        # Confirm order with edits and delivery partner
        enhanced_order_service = get_enhanced_order_service()
        result = enhanced_order_service.confirm_order_by_distributor(order_id, user.id, item_edits, delivery_partner_id)
        
        return jsonify(result), 200 if result['success'] else 400
        
    except Exception as e:
        logger.error(f"Error confirming order: {str(e)}")
        return jsonify({'error': 'Error confirming order'}), 500

@chatbot_bp.route('/search_products', methods=['POST'])
def search_products():
    """Search for products in Azure AI Search"""
    try:
        from app.input_validation import sanitize_string, MAX_LENGTHS
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        search_query_raw = data.get('query', '').strip()
        if not search_query_raw:
            return jsonify({'error': 'Search query is required'}), 400
        
        search_query = sanitize_string(search_query_raw, max_length=MAX_LENGTHS['search_query'])
        if not search_query:
            return jsonify({'error': 'Invalid search query'}), 400
        
        search_service = get_search_service()
        if not search_service.is_available():
            return jsonify({'error': 'Azure AI Search is not configured'}), 503
        
        # Search for products
        results = search_service.search_products(search_query, top=20)
        
        product_list = []
        for product in results:
            product_name = product.get('product_name') or product.get('name') or product.get('title') or 'Unknown Product'
            product_list.append({
                'id': product.get('id') or product.get('product_id') or product_name,
                'name': product_name,
                'description': product.get('description') or product.get('content') or product.get('text') or '',
                'full_data': product  # Include full data for detailed view
            })
        
        return jsonify({
            'products': product_list,
            'count': len(product_list)
        }), 200
        
    except Exception as e:
        logger.error(f"Error searching products: {str(e)}")
        return jsonify({'error': f'Error searching products: {str(e)}'}), 500

@chatbot_bp.route('/get_product_details', methods=['POST'])
def get_product_details():
    """Get detailed information about a specific product"""
    try:
        from app.input_validation import sanitize_string, MAX_LENGTHS
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        product_id_raw = data.get('product_id')
        product_name_raw = data.get('product_name')
        
        # Validate that at least one is provided
        if not product_id_raw and not product_name_raw:
            return jsonify({'error': 'Product ID or name is required'}), 400
        
        # Sanitize product_id if provided - can be string or integer
        product_id = None
        product_id_str = None
        if product_id_raw:
            # Try to convert to int if possible, but also keep as string for search
            try:
                product_id = int(product_id_raw)
                if product_id <= 0:
                    product_id = None
                else:
                    product_id_str = str(product_id_raw)  # Keep original as string for search
            except (ValueError, TypeError):
                # If not an integer, treat as string (could be product name or string ID)
                product_id_str = str(product_id_raw).strip()
                if not product_id_str:
                    product_id = None
                    product_id_str = None
        
        # Sanitize product_name if provided
        product_name = None
        if product_name_raw:
            product_name = sanitize_string(str(product_name_raw), max_length=MAX_LENGTHS['product_name'])
            if not product_name:
                return jsonify({'error': 'Invalid product name'}), 400
        
        search_service = get_search_service()
        if not search_service.is_available():
            return jsonify({'error': 'Azure AI Search is not configured'}), 503
        
        # Try to get product by name first, then by ID/string
        product = None
        if product_name:
            product = search_service.get_product_by_name(product_name)
        
        if not product and product_id_str:
            # Search by ID string (could be integer ID or product name)
            results = search_service.search_products(product_id_str, top=10)
            if results:
                # If we have a numeric ID, try to find exact match first
                if product_id:
                    for p in results:
                        p_id = p.get('id') or p.get('product_id')
                        if p_id and (str(p_id) == str(product_id) or int(p_id) == product_id):
                            product = p
                            break
                # If no exact match or not numeric, use first result
                if not product:
                    product = results[0]
        
        if not product:
            return jsonify({'error': 'Product not found'}), 404
        
        # Extract structured information from the product document
        product_info = {
            'id': product.get('id') or product.get('product_id'),
            'name': product.get('product_name') or product.get('name') or product.get('title') or 'Unknown Product',
            'description': product.get('description') or product.get('content') or product.get('text') or '',
            'generic_name': product.get('generic_name') or product.get('composition') or '',
            'therapeutic_class': product.get('therapeutic_class') or product.get('class') or '',
            'key_uses': product.get('key_uses') or product.get('uses') or product.get('indications') or '',
            'mechanism_of_action': product.get('mechanism_of_action') or product.get('moa') or '',
            'dosage': product.get('dosage') or product.get('administration') or product.get('dosage_administration') or '',
            'safety_profile': product.get('safety_profile') or product.get('side_effects') or product.get('adverse_effects') or '',
            'pack_size': product.get('pack_size') or product.get('packaging') or '',
            'full_content': product.get('content') or product.get('text') or product.get('description') or ''
        }
        
        return jsonify({
            'product': product_info
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting product details: {str(e)}", exc_info=True)
        return jsonify({'error': f'Error getting product details: {str(e)}'}), 500

@chatbot_bp.route('/cancel_order_action', methods=['POST'])
def cancel_order_action():
    """Cancel order by MR"""
    try:
        from app.input_validation import validate_and_sanitize_order_id
        
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'mr':
            return jsonify({'error': 'Only MRs can cancel their orders'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        order_id_raw = data.get('order_id')
        if not order_id_raw:
            return jsonify({'error': 'Order ID is required'}), 400
        
        order_id = validate_and_sanitize_order_id(order_id_raw)
        if not order_id:
            return jsonify({'error': 'Invalid order ID format'}), 400
        
        # Cancel order
        enhanced_order_service = get_enhanced_order_service()
        result = enhanced_order_service.cancel_order_by_mr(order_id, user.id)
        
        return jsonify(result), 200 if result['success'] else 400
        
    except Exception as e:
        logger.error(f"Error cancelling order: {str(e)}")
        return jsonify({'error': 'Error cancelling order'}), 500

# Bulk order endpoints removed - bulk order functionality has been removed

# Export orders endpoint removed - was part of bulk order functionality
# @chatbot_bp.route('/export_orders_excel', methods=['POST'])
def export_orders_excel_removed():
    """Export orders to Excel format"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        data = request.get_json()
        order_ids = data.get('order_ids')  # None = all orders
        
        from app.models import Order, OrderItem
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        # Query orders
        if user.role == 'distributor':
            # Get MRs in distributor's area
            mr_ids = [u.id for u in User.query.filter_by(role='mr', area=user.area).all()]
            if order_ids:
                orders = Order.query.filter(Order.order_id.in_(order_ids), Order.mr_id.in_(mr_ids)).all()
            else:
                orders = Order.query.filter(Order.mr_id.in_(mr_ids)).all()
        elif user.role == 'mr':
            if order_ids:
                orders = Order.query.filter(Order.order_id.in_(order_ids), Order.mr_id == user_id).all()
            else:
                orders = Order.query.filter_by(mr_id=user_id).all()
        else:
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Orders Export"
        
        # Header row
        headers = ['Order ID', 'Date', 'Status', 'MR Name', 'Customer', 'Total Items', 'Subtotal', 'Tax', 'Grand Total (MMK)']
        header_fill = PatternFill(start_color="2563eb", end_color="2563eb", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Data rows
        for row_num, order in enumerate(orders, 2):
            ws.cell(row=row_num, column=1, value=order.order_id)
            ws.cell(row=row_num, column=2, value=order.created_at.strftime('%Y-%m-%d %H:%M:%S') if order.created_at else 'N/A')
            ws.cell(row=row_num, column=3, value=(order.status or 'N/A').replace('_', ' ').title())
            ws.cell(row=row_num, column=4, value=order.mr.name if order.mr else 'N/A')
            ws.cell(row=row_num, column=5, value=order.customer.name if order.customer else 'N/A')
            
            # Calculate totals
            total_items = sum(item.quantity + (item.free_quantity or 0) for item in order.order_items)
            subtotal = float(order.subtotal) if order.subtotal else 0.0
            tax = float(order.tax_amount) if order.tax_amount else 0.0
            grand_total = float(order.total_amount) if order.total_amount else 0.0
            
            ws.cell(row=row_num, column=6, value=total_items)
            ws.cell(row=row_num, column=7, value=subtotal)
            ws.cell(row=row_num, column=8, value=tax)
            ws.cell(row=row_num, column=9, value=grand_total)
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[col_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename=orders_export_{datetime.now().strftime("%Y%m%d")}.xlsx'}
        )
        
    except Exception as e:
        logger.error(f"Error exporting orders to Excel: {str(e)}")
        return jsonify({'error': 'Error exporting to Excel'}), 500

@chatbot_bp.route('/advanced_search', methods=['POST'])
def advanced_search():
    """Advanced search across orders, products, and customers"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        from app.input_validation import sanitize_string, MAX_LENGTHS, sanitize_dict
        
        query_raw = data.get('query', '').strip().lower()
        query = sanitize_string(query_raw, max_length=MAX_LENGTHS['search_query']) if query_raw else ''
        
        filters_raw = data.get('filters', {})
        filters = sanitize_dict(filters_raw) if isinstance(filters_raw, dict) else {}
        
        results = {
            'orders': [],
            'products': [],
            'customers': []
        }
        
        # Search orders
        if user.role == 'mr':
            orders = Order.query.filter_by(mr_id=user_id).all()
        elif user.role == 'distributor':
            mr_ids = [u.id for u in User.query.filter_by(role='mr', area=user.area).all()]
            orders = Order.query.filter(Order.mr_id.in_(mr_ids)).all()
        else:
            orders = []
        
        # Filter orders by query
        if query:
            filtered_orders = []
            for order in orders:
                if (query in order.order_id.lower() or
                    (order.customer and query in order.customer.name.lower()) or
                    (order.mr and query in order.mr.name.lower())):
                    filtered_orders.append(order)
            results['orders'] = [{
                'order_id': o.order_id,
                'status': o.status,
                'total_amount': float(o.total_amount) if o.total_amount else 0.0,
                'order_date': o.created_at.strftime('%Y-%m-%d') if o.created_at else 'N/A',
                'customer_name': o.customer.name if o.customer else None
            } for o in filtered_orders[:20]]  # Limit to 20
        
        # Search products
        if user.role == 'mr':
            db_service = get_db_service()
            products = db_service.get_products_from_dealer_stock(user.area)
        else:
            products = Product.query.all()
        
        if query:
            filtered_products = [p for p in products if query in (p.product_name or '').lower()]
            results['products'] = [{
                'product_name': p.product_name,
                'product_code': getattr(p, 'product_code', None),
                'price': float(p.price) if p.price else 0.0
            } for p in filtered_products[:20]]
        
        # Search customers (for MRs)
        if user.role == 'mr' and query:
            customers = Customer.query.filter_by(mr_id=user_id).all()
            filtered_customers = [c for c in customers if query in c.name.lower() or query in c.unique_id.lower()]
            results['customers'] = [{
                'name': c.name,
                'unique_id': c.unique_id,
                'email': c.email,
                'phone': c.phone
            } for c in filtered_customers[:10]]
        
        return jsonify({
            'success': True,
            'results': results,
            'orders': results['orders'],
            'products': results['products']
        }), 200
        
    except Exception as e:
        logger.error(f"Error in advanced search: {str(e)}")
        return jsonify({'success': False, 'error': 'Error performing search'}), 500

@chatbot_bp.route('/reject_order_action', methods=['POST'])
def reject_order_action():
    """Reject order by distributor"""
    try:
        from app.input_validation import validate_and_sanitize_order_id, sanitize_string, MAX_LENGTHS
        
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Only distributors can reject orders'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        order_id_raw = data.get('order_id')
        if not order_id_raw:
            return jsonify({'error': 'Order ID is required'}), 400
        
        order_id = validate_and_sanitize_order_id(order_id_raw)
        if not order_id:
            return jsonify({'error': 'Invalid order ID format'}), 400
        
        # Sanitize rejection_reason if provided
        rejection_reason_raw = data.get('reason', 'No reason provided')
        rejection_reason = sanitize_string(
            str(rejection_reason_raw), 
            max_length=MAX_LENGTHS['rejection_reason']
        ) if rejection_reason_raw else 'No reason provided'
        
        # Reject order
        enhanced_order_service = get_enhanced_order_service()
        result = enhanced_order_service.reject_order_by_distributor(order_id, user.id, rejection_reason)
        
        return jsonify(result), 200 if result['success'] else 400
        
    except Exception as e:
        logger.error(f"Error rejecting order: {str(e)}")
        return jsonify({'error': 'Error rejecting order'}), 500

@chatbot_bp.route('/add_customer', methods=['POST'])
def add_customer():
    """Add new customer for MR"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'mr':
            return jsonify({'error': 'Only MRs can add customers'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        from app.input_validation import (
            sanitize_string, validate_email, validate_phone, MAX_LENGTHS
        )
        
        # Validate and sanitize inputs
        name_raw = data.get('name', '').strip()
        name = sanitize_string(name_raw, max_length=MAX_LENGTHS['customer_name'])
        if not name:
            return jsonify({'error': 'Customer name is required and must be valid'}), 400
        
        email_raw = data.get('email', '').strip()
        email = None
        if email_raw:
            email = sanitize_string(email_raw, max_length=MAX_LENGTHS['email'])
            if email and not validate_email(email):
                return jsonify({'error': 'Invalid email format'}), 400
        
        phone_raw = data.get('phone', '').strip()
        phone = None
        if phone_raw:
            phone = sanitize_string(phone_raw, max_length=MAX_LENGTHS['phone'])
            if phone and not validate_phone(phone):
                return jsonify({'error': 'Invalid phone number format'}), 400
        
        address_raw = data.get('address', '').strip()
        address = sanitize_string(address_raw, max_length=500) if address_raw else None
        
        # Create new customer
        customer = Customer(
            name=name,
            email=email if email else None,
            phone=phone if phone else None,
            address=address if address else None,
            mr_unique_id=user.unique_id,
            mr_id=user.id,
            is_active=True
        )
        
        # Generate unique ID
        customer.generate_unique_id()
        
        # Save to database
        try:
            db.session.add(customer)
            db.session.commit()
        except Exception as e:
            logger.error(f"Error saving customer to database: {str(e)}")
            db.session.rollback()
            raise
        
        # Store selected customer in session
        session['selected_customer_id'] = customer.id
        session['selected_customer_unique_id'] = customer.unique_id
        
        # Get available products for the MR
        db_service = get_db_service()
        if user.area:
            products = db_service.get_products_from_dealer_stock(user.area)
        else:
            from app.models import Product
            products = Product.query.all()
        
        # Build product list for interactive UI with FOC information using helper
        pricing_service = get_pricing_service()
        product_list = build_product_list_with_foc(products, pricing_service)
        
        # Create response message with customer info
        response = f"• Great! I can help you place an order.\n• **Ordering for:** {customer.name} ({customer.unique_id})\n• Please use the product selection form below to select products and quantities."
        
        # Save conversation
        save_conversation(user.id, f"Added new customer: {customer.name}", response)
        
        return jsonify({
            'success': True,
            'message': f'Customer {customer.name} added successfully',
            'customer': {
                'id': customer.id,
                'unique_id': customer.unique_id,
                'name': customer.name,
                'email': customer.email,
                'phone': customer.phone,
                'address': customer.address
            },
            'response': response,
            'products': product_list,
            'interactive_product_selection': True,  # Flag for frontend to render interactive UI
            'show_product_table': True,  # Flag to show table outside selection box
            'action_buttons': [{'text': 'Change Customer', 'action': 'change_customer'}]
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Error adding customer: {str(e)}")
        return jsonify({'error': f'Error adding customer: {str(e)}'}), 500

@chatbot_bp.route('/place_order', methods=['POST'])
def place_order():
    """Place order from cart"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        enhanced_order_service = get_enhanced_order_service()
        
        # Get customer details if MR
        user = User.query.get(user_id)
        customer_id = None
        if user and user.role == 'mr':
            # For MRs, customer selection is mandatory
            if 'selected_customer_id' not in session:
                return jsonify({
                    'success': False,
                    'message': 'Please select a customer before placing an order. Use "Place Order" to select a customer first.',
                    'requires_customer_selection': True
                }), 400
            customer_id = session.get('selected_customer_id')
        
        result = enhanced_order_service.place_order(user_id, customer_id=customer_id)
        
        if result['success']:
            # Clear selected customer after order is placed (MR can place order for one customer at a time)
            if user and user.role == 'mr' and 'selected_customer_id' in session:
                session.pop('selected_customer_id', None)
                session.pop('selected_customer_unique_id', None)
            
            return jsonify({
                'success': True,
                'message': result['message'],
                'order_id': result['order_id'],
                'order_summary': result['order_summary']
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        logger.error(f"Error placing order: {str(e)}")
        return jsonify({'error': 'Error placing order'}), 500

@chatbot_bp.route('/cart/add', methods=['POST'])
def add_to_cart_api():
    """Add product to cart via API (for bulk additions)"""
    try:
        from app.input_validation import (
            validate_and_sanitize_product_code, validate_quantity
        )
        
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        # Validate and sanitize product_code
        product_code_raw = data.get('product_code')
        if not product_code_raw:
            return jsonify({'error': 'Product code is required'}), 400
        
        product_code = validate_and_sanitize_product_code(product_code_raw)
        if not product_code:
            return jsonify({'error': 'Invalid product code format'}), 400
        
        # Validate quantity
        quantity_raw = data.get('quantity')
        if not quantity_raw:
            return jsonify({'error': 'Quantity is required'}), 400
        
        if not validate_quantity(quantity_raw):
            return jsonify({'error': 'Quantity must be a positive integer'}), 400
        
        quantity = int(quantity_raw)
        
        # Get product from dealer stock
        db_service = get_db_service()
        products = db_service.get_products_from_dealer_stock(user.area)
        
        product = None
        for p in products:
            if p.get('product_code') == product_code:
                product = p
                break
        
        if not product:
            return jsonify({
                'success': False,
                'error': f'Product with code "{product_code}" not found in your area. The product may no longer be available or the product code may have changed.'
            }), 200  # Return 200 with success: false so frontend can parse it
        
        # Get pricing
        pricing_service = get_pricing_service()
        pricing = pricing_service.calculate_product_pricing(
            product.get('product_id'),
            quantity
        )
        
        if 'error' in pricing:
            return jsonify({
                'success': False,
                'error': pricing['error']
            }), 200  # Return 200 with success: false so frontend can parse it
        
        # Add to cart
        unit_price = pricing['pricing']['final_price']
        logger.info(f"Adding to cart: product_code={product_code}, product_id={product.get('product_id')}, quantity={quantity}, unit_price={unit_price}")
        
        try:
            cart_item, message = db_service.add_to_cart(
                user_id,
                product.get('product_id'),
                product_code,
                product.get('product_name'),
                quantity,
                unit_price
            )
            
            if cart_item:
                logger.info(f"Successfully added to cart: cart_item_id={cart_item.id if hasattr(cart_item, 'id') else 'N/A'}")
                return jsonify({
                    'success': True,
                    'message': message,
                    'cart_item': {
                        'id': cart_item.id,
                        'product_code': cart_item.product_code,
                        'product_name': cart_item.product_name,
                        'quantity': cart_item.quantity,
                        'unit_price': cart_item.unit_price
                    },
                    'pricing': pricing
                }), 200
            else:
                logger.warning(f"add_to_cart returned None for product_code={product_code}")
                return jsonify({
                    'success': False,
                    'error': message or 'Failed to add product to cart'
                }), 200  # Return 200 with success: false so frontend can parse it
        except Exception as db_error:
            logger.error(f"Database error adding to cart: {str(db_error)}")
            import traceback
            traceback.print_exc()
            return jsonify({
                'success': False,
                'error': f'Database error: {str(db_error)}'
            }), 200  # Return 200 with success: false so frontend can parse it
            
    except Exception as e:
        logger.error(f"Error adding to cart: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'success': False,
            'error': f'Error adding product to cart: {str(e)}'
        }), 200  # Return 200 with success: false so frontend can parse it

@chatbot_bp.route('/cart', methods=['GET'])
def get_cart():
    """Get user's cart items"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        db_service = get_db_service()
        cart_items = db_service.get_cart_items(user_id)
        cart_data = []
        
        # Get pricing service to calculate FOC for cart items
        from app.pricing_service import PricingService
        pricing_service = PricingService()
        
        for item in cart_items:
            # Get product name from cart item (now stored in CartItem) or fallback
            product_name = item.product_code
            if hasattr(item, 'product_name') and item.product_name:
                product_name = item.product_name
            elif hasattr(item, 'product') and item.product:
                product_name = item.product.product_name
            
            # Calculate current pricing with FOC
            if hasattr(item, 'product') and item.product:
                pricing_result = pricing_service.calculate_product_pricing(item.product.id, item.quantity)
                if 'error' not in pricing_result:
                    scheme_info = pricing_result.get('scheme', {})
                    pricing_info = pricing_result.get('pricing', {})
                    free_quantity = scheme_info.get('free_quantity', 0)
                    paid_quantity = scheme_info.get('paid_quantity', item.quantity)
                    total_quantity = scheme_info.get('total_quantity', item.quantity)
                    scheme_name = scheme_info.get('name')
                    scheme_applied = scheme_info.get('applied', False)
                    total_price = pricing_info.get('total_amount', item.total_price)
                    final_price = pricing_info.get('final_price', item.unit_price)
                else:
                    # Fallback to cart item values
                    free_quantity = getattr(item, 'free_quantity', 0)
                    paid_quantity = item.quantity
                    total_quantity = item.quantity
                    scheme_name = None
                    scheme_applied = False
                    total_price = item.total_price
                    final_price = item.unit_price
            else:
                # Fallback to cart item values
                free_quantity = getattr(item, 'free_quantity', 0)
                paid_quantity = item.quantity
                total_quantity = item.quantity
                scheme_name = None
                scheme_applied = False
                total_price = item.total_price
                final_price = item.unit_price
            
            cart_data.append({
                'id': item.id,
                'product_code': item.product_code,
                'product_name': product_name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': total_price,
                'base_price': getattr(item, 'base_price', item.unit_price),
                'discount_amount': getattr(item, 'discount_amount', 0),
                'final_price': final_price,
                'scheme_applied': scheme_applied,
                'scheme_name': scheme_name,
                'free_quantity': free_quantity,
                'paid_quantity': paid_quantity,
                'total_quantity': total_quantity
            })
        
        # Calculate subtotal, tax, and grand total
        from flask import current_app
        subtotal = sum(item['total_price'] for item in cart_data)
        tax_rate = current_app.config.get('TAX_RATE', 0.05)  # Get from config, default 5%
        tax_amount = subtotal * tax_rate
        grand_total = subtotal + tax_amount
        
        # Get customer info if available (for MRs)
        customer_id = None
        customer_name = ''
        if 'selected_customer_id' in session:
            customer_id = session.get('selected_customer_id')
            from app.models import Customer
            customer = Customer.query.get(customer_id)
            if customer:
                customer_name = customer.name
        
        return jsonify({
            'cart_items': cart_data,
            'subtotal': round(subtotal, 2),
            'tax_rate': tax_rate,
            'tax_amount': round(tax_amount, 2),
            'grand_total': round(grand_total, 2),
            'customer_id': customer_id,
            'customer_name': customer_name
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting cart: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'Error getting cart: {str(e)}'}), 500

@chatbot_bp.route('/cart/<int:item_id>', methods=['DELETE'])
def remove_from_cart(item_id):
    """Remove item from cart"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        db_service = get_db_service()
        
        # Get cart item before removing to unblock quantities for MR orders
        cart_item = CartItem.query.get(item_id)
        if cart_item:
            user = User.query.get(user_id)
            # Unblock quantities for MR orders when removing from cart
            if user and user.role == 'mr' and user.area and cart_item:
                enhanced_order_service = get_enhanced_order_service()
                enhanced_order_service._unblock_quantity_for_mr_order(
                    user=user,
                    product_code=cart_item.product_code,
                    quantity=cart_item.quantity
                )
                logger.info(f"Unblocked {cart_item.quantity} units of {cart_item.product_code} for MR order (removed from cart)")
        
        success, message = db_service.remove_from_cart(item_id)
        
        if success:
            # Return updated cart items (even if empty)
            cart_items = db_service.get_cart_items(user_id)
            cart_data = []
            for item in cart_items:
                # Get product name from cart item (now stored in CartItem) or fallback
                product_name = item.product_code
                if hasattr(item, 'product_name') and item.product_name:
                    product_name = item.product_name
                elif hasattr(item, 'product') and item.product:
                    product_name = item.product.product_name
                
                cart_data.append({
                    'id': item.id,
                    'product_code': item.product_code,
                    'product_name': product_name,
                    'quantity': item.quantity,
                    'unit_price': item.unit_price,
                    'total_price': item.total_price,
                    'base_price': getattr(item, 'base_price', item.unit_price),
                    'discount_amount': getattr(item, 'discount_amount', 0),
                    'final_price': getattr(item, 'final_price', item.unit_price),
                    'scheme_applied': getattr(item, 'scheme_applied', None),
                    'free_quantity': getattr(item, 'free_quantity', 0),
                    'paid_quantity': getattr(item, 'paid_quantity', item.quantity)
                })
            
            # Calculate totals
            subtotal = sum(item.get('total_price', 0) for item in cart_data)
            from flask import current_app
            tax_rate = current_app.config.get('TAX_RATE', 0.05)  # Get from config, default 5%
            tax_amount = subtotal * tax_rate
            grand_total = subtotal + tax_amount
            
            return jsonify({
                'message': message, 
                'cart_items': cart_data,
                'subtotal': subtotal,
                'tax_amount': tax_amount,
                'grand_total': grand_total
            }), 200
        else:
            return jsonify({'error': message}), 400
            
    except Exception as e:
        logger.error(f"Error removing from cart: {str(e)}")
        return jsonify({'error': 'Error removing from cart'}), 500

@chatbot_bp.route('/cart/<int:item_id>/quantity', methods=['PATCH'])
def update_cart_quantity(item_id):
    """Update cart item quantity"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        data = request.get_json()
        new_quantity = data.get('quantity')
        
        if not new_quantity or new_quantity < 1:
            return jsonify({'error': 'Invalid quantity'}), 400
        
        db_service = get_db_service()
        
        # Get cart item
        cart_item = CartItem.query.filter_by(id=item_id, user_id=user_id).first()
        if not cart_item:
            return jsonify({'error': 'Cart item not found'}), 404
        
        # Get product to check availability
        product = Product.query.get(cart_item.product_id)
        if not product:
            return jsonify({'error': 'Product not found'}), 404
        
        # Check stock availability - for MR users, get from dealer stock
        user = User.query.get(user_id)
        available_quantity = 0
        if user and user.role == 'mr' and user.area:
            # Get available quantity from dealer stock
            from app.models import DealerWiseStockDetails
            from datetime import date
            today = date.today()
            dealers_in_area = User.query.filter_by(role='distributor', area=user.area).all()
            if dealers_in_area:
                stock_details = DealerWiseStockDetails.query.filter(
                    DealerWiseStockDetails.product_code == cart_item.product_code,
                    DealerWiseStockDetails.status == 'confirmed',
                    DealerWiseStockDetails.dealer_unique_id.in_([d.unique_id for d in dealers_in_area]),
                    DealerWiseStockDetails.available_for_sale > 0
                ).filter(
                    db.or_(
                        DealerWiseStockDetails.expiry_date >= today,
                        DealerWiseStockDetails.expiry_date.is_(None)
                    )
                ).all()
                available_quantity = sum(s.available_for_sale for s in stock_details)
        else:
            # For non-MR users, use product available_for_sale if it exists
            available_quantity = getattr(product, 'available_for_sale', 999999)
        
        if new_quantity > available_quantity:
            return jsonify({'error': f'Insufficient stock. Only {available_quantity} units available.'}), 400
        
        # For MR orders, adjust blocked quantities when quantity changes
        old_quantity = cart_item.quantity
        quantity_diff = new_quantity - old_quantity
        
        if user and user.role == 'mr' and user.area:
            enhanced_order_service = get_enhanced_order_service()
            if quantity_diff > 0:
                # Increase - block more
                enhanced_order_service._block_quantity_for_mr_order(
                    user=user,
                    product_code=cart_item.product_code,
                    quantity=quantity_diff
                )
                logger.info(f"Blocked additional {quantity_diff} units of {cart_item.product_code} for MR order (quantity increased)")
            elif quantity_diff < 0:
                # Decrease - unblock some
                enhanced_order_service._unblock_quantity_for_mr_order(
                    user=user,
                    product_code=cart_item.product_code,
                    quantity=abs(quantity_diff)
                )
                logger.info(f"Unblocked {abs(quantity_diff)} units of {cart_item.product_code} for MR order (quantity decreased)")
        
        # Update quantity
        cart_item.quantity = new_quantity
        cart_item.paid_quantity = new_quantity
        
        # Recalculate pricing
        pricing_service = get_pricing_service()
        pricing = pricing_service.calculate_product_pricing(product.id, new_quantity)
        
        if 'error' in pricing:
            return jsonify({'error': pricing['error']}), 400
        
        # Update cart item pricing
        cart_item.base_price = pricing['base_price']
        cart_item.discount_amount = pricing['discount']['amount']
        cart_item.final_price = pricing['pricing']['final_price']
        cart_item.free_quantity = pricing['scheme']['free_quantity']
        cart_item.paid_quantity = pricing['scheme']['paid_quantity']
        cart_item.total_price = cart_item.final_price * cart_item.paid_quantity
        
        db.session.commit()
        
        # Return updated cart items
        cart_items = db_service.get_cart_items(user_id)
        cart_data = []
        for item in cart_items:
            # Get product name from cart item (now stored in CartItem) or fallback
            product_name = item.product_code
            if hasattr(item, 'product_name') and item.product_name:
                product_name = item.product_name
            elif hasattr(item, 'product') and item.product:
                product_name = item.product.product_name
            
            cart_data.append({
                'id': item.id,
                'product_code': item.product_code,
                'product_name': product_name,
                'quantity': item.quantity,
                'unit_price': item.unit_price,
                'total_price': item.total_price,
                'base_price': getattr(item, 'base_price', item.unit_price),
                'discount_amount': getattr(item, 'discount_amount', 0),
                'final_price': getattr(item, 'final_price', item.unit_price),
                'scheme_applied': getattr(item, 'scheme_applied', None),
                'free_quantity': getattr(item, 'free_quantity', 0),
                'paid_quantity': getattr(item, 'paid_quantity', item.quantity)
            })
        
        # Calculate totals
        subtotal = sum(item.get('total_price', 0) for item in cart_data)
        tax_amount = subtotal * 0.05  # 5% tax
        grand_total = subtotal + tax_amount
        
        return jsonify({
            'message': 'Quantity updated successfully', 
            'cart_items': cart_data,
            'subtotal': subtotal,
            'tax_amount': tax_amount,
            'grand_total': grand_total
        }), 200
        
    except Exception as e:
        logger.error(f"Error updating cart quantity: {str(e)}")
        db.session.rollback()
        return jsonify({'error': 'Error updating cart quantity'}), 500

@chatbot_bp.route('/api/quick-stats', methods=['GET'])
def get_quick_stats():
    """Get quick statistics for the current user"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'success': False, 'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 404
        
        stats = {
            'pendingOrders': 0,
            'totalOrders': 0,
            'cartItems': 0
        }
        
        if user.role == 'mr':
            # Count pending orders for this MR
            from app.models import Order
            stats['pendingOrders'] = Order.query.filter_by(
                mr_id=user_id,
                status='pending'
            ).count()
            stats['totalOrders'] = Order.query.filter_by(mr_id=user_id).count()
            
            # Count cart items
            from app.models import CartItem
            stats['cartItems'] = CartItem.query.filter_by(user_id=user_id).count()
            
        elif user.role == 'distributor':
            # Count pending orders in distributor's area
            from app.models import Order, User as UserModel
            mr_ids_in_area = [u.id for u in UserModel.query.filter_by(role='mr', area=user.area).all()]
            if mr_ids_in_area:
                stats['pendingOrders'] = Order.query.filter(
                    Order.mr_id.in_(mr_ids_in_area),
                    Order.status == 'pending'
                ).count()
                stats['totalOrders'] = Order.query.filter(Order.mr_id.in_(mr_ids_in_area)).count()
        
        return jsonify({
            'success': True,
            'stats': stats
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting quick stats: {str(e)}")
        return jsonify({'success': False, 'error': 'Error getting stats'}), 500

@chatbot_bp.route('/api/products', methods=['GET'])
def get_products_api():
    """Get products API for autocomplete"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        db_service = get_db_service()
        user = User.query.get(user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        # Get products - For MRs, get from dealer_wise_stock_details; for distributors, use Product table
        area = user.area if user else None
        
        if user.role == 'mr' and area:
            products = db_service.get_products_from_dealer_stock(area)
        else:
            warehouse = db_service.get_warehouse_by_area(area) if area else None
            if warehouse:
                products = db_service.get_products_by_warehouse(warehouse.id)
            else:
                products = Product.query.all()
        
        product_list = []
        for product in products:
            # Handle both Product objects and dictionaries from dealer stock
            if isinstance(product, dict):
                # Already a dictionary from dealer stock
                product_list.append({
                    'product_code': product.get('product_code', ''),
                    'product_name': product.get('product_name', ''),
                    'product_description': '',
                    'price_of_product': float(product.get('sales_price', product.get('price', 0))),
                    'available_for_sale': int(product.get('available_quantity', 0)),
                    'sales_price': float(product.get('sales_price', 0)),
                    'batch_number': '',
                    'expiry_date': product.get('earliest_expiry')
                })
            else:
                # Product object from database
                product_list.append({
                    'product_code': str(product.id),  # Use ID as code for new schema
                    'product_name': product.product_name,
                    'product_description': '',
                    'price_of_product': float(product.price) if product.price else 0.0,
                    'available_for_sale': 0,  # Will be calculated from dealer stock
                    'sales_price': float(product.price) if product.price else 0.0,
                    'batch_number': '',
                    'expiry_date': None
                })
        
        return jsonify({'products': product_list}), 200
        
    except Exception as e:
        logger.error(f"Error getting products: {str(e)}")
        return jsonify({'error': 'Error getting products'}), 500

@chatbot_bp.route('/distributor/confirm_order', methods=['POST'])
def distributor_confirm_order():
    """Distributor confirm order endpoint"""
    try:
        data = request.get_json()
        order_id = data.get('order_id')
        distributor_user_id = session.get('user_id')
        
        if not distributor_user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(distributor_user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Access denied. Distributor access required.'}), 403
        
        # Get delivery partner ID from request if provided
        delivery_partner_id = data.get('delivery_partner_id')
        if delivery_partner_id:
            try:
                delivery_partner_id = int(delivery_partner_id)
            except (ValueError, TypeError):
                delivery_partner_id = None
        
        enhanced_order_service = get_enhanced_order_service()
        result = enhanced_order_service.confirm_order_by_distributor(order_id, distributor_user_id, item_edits=None, delivery_partner_id=delivery_partner_id)
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'invoice_number': result['invoice_number']
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        logger.error(f"Error confirming order: {str(e)}")
        return jsonify({'error': 'Error confirming order'}), 500

@chatbot_bp.route('/stock/pending', methods=['GET'])
def get_pending_stock_arrivals():
    """Get pending stock arrivals for dealer"""
    try:
        dealer_user_id = session.get('user_id')
        if not dealer_user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(dealer_user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Access denied. Distributor access required.'}), 403
        
        stock_service = get_stock_management_service()
        result = stock_service.get_pending_stock_arrivals(user.unique_id)
        
        if result['success']:
            return jsonify({
                'success': True,
                'stocks': result['stocks'],
                'count': result['count']
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        logger.error(f"Error getting pending stock arrivals: {str(e)}")
        return jsonify({'error': 'Error getting pending stock arrivals'}), 500

@chatbot_bp.route('/stock/confirm', methods=['POST'])
def confirm_stock_arrival():
    """Dealer confirm stock arrival"""
    try:
        dealer_user_id = session.get('user_id')
        if not dealer_user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(dealer_user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Access denied. Distributor access required.'}), 403
        
        data = request.get_json()
        stock_detail_id = data.get('stock_detail_id')
        received_quantity = data.get('received_quantity')
        adjustment_reason = data.get('adjustment_reason')
        
        stock_service = get_stock_management_service()
        result = stock_service.confirm_stock_arrival(
            stock_detail_id,
            dealer_user_id,
            received_quantity,
            adjustment_reason
        )
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'stock_detail': result['stock_detail'],
                'quantity_adjusted': result.get('quantity_adjusted', False)
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        logger.error(f"Error confirming stock arrival: {str(e)}")
        return jsonify({'error': 'Error confirming stock arrival'}), 500

@chatbot_bp.route('/stock/adjust', methods=['POST'])
def adjust_stock_quantity():
    """Adjust stock quantity after confirmation"""
    try:
        dealer_user_id = session.get('user_id')
        if not dealer_user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(dealer_user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Access denied. Distributor access required.'}), 403
        
        data = request.get_json()
        stock_detail_id = data.get('stock_detail_id')
        new_quantity = data.get('new_quantity')
        reason = data.get('reason', 'Quantity adjustment')
        
        if not new_quantity or new_quantity < 0:
            return jsonify({'error': 'Invalid quantity'}), 400
        
        stock_service = get_stock_management_service()
        result = stock_service.adjust_stock_quantity(
            stock_detail_id,
            dealer_user_id,
            new_quantity,
            reason
        )
        
        if result['success']:
            return jsonify({
                'success': True,
                'message': result['message'],
                'stock_detail': result['stock_detail']
            }), 200
        else:
            return jsonify({
                'success': False,
                'message': result['message']
            }), 400
            
    except Exception as e:
        logger.error(f"Error adjusting stock quantity: {str(e)}")
        return jsonify({'error': 'Error adjusting stock quantity'}), 500


# ============= COMPANY REPORT SYSTEM =============

def handle_company_requests(user_message, user):
    """Handle all company user requests for report generation"""
    try:
        from app.company_report_service import CompanyReportService
        report_service = CompanyReportService()
        
        message_lower = user_message.lower().strip()
        
        # Check for report generation keywords
        report_keywords = ['generate report', 'export data', 'download report', 'get report', 
                          'export table', 'i want a report', 'i need a report']
        
        if any(keyword in message_lower for keyword in report_keywords) or message_lower == 'generate report':
            # Show table selection
            available_tables = report_service.get_available_tables()
            
            response = f"""**📊 Database Report Generation**

Welcome {user.name}! I can help you export data from any table in our database.

**Available Tables:**

"""
            
            tables_list = []
            for table_key, table_info in available_tables.items():
                response += f"• **{table_info['name']}** (`{table_key}`)\n"
                response += f"  _{len(table_info['columns'])} columns available_\n\n"
                tables_list.append({
                    'key': table_key,
                    'name': table_info['name'],
                    'columns': table_info['columns']
                })
            
            response += "**How it works:**\n"
            response += "1. Select a table from the list above\n"
            response += "2. Choose which columns to include (or select all)\n"
            response += "3. Receive the CSV file via email\n\n"
            response += "Please select a table to continue."
            
            save_conversation(user.id, user_message, response)
            
            return jsonify({
                'response': response,
                'interactive_report_selection': True,
                'tables': tables_list,
                'action_buttons': []
            }), 200
        
        # Greeting for company users
        elif any(greeting in message_lower for greeting in ['hi', 'hello', 'hey', 'good morning', 'good afternoon']):
            response = f"""Hello {user.name}! 👋

Welcome to HV Company Analytics.

I'm here to help you generate database reports and export data. You can request reports from any table in our system.

What would you like to do?"""
            
            save_conversation(user.id, user_message, response)
            
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Generate Report', 'action': 'generate_report'},
                    {'text': 'Help', 'action': 'help'}
                ]
            }), 200
        
        # Help message
        elif 'help' in message_lower:
            response = """**🆘 Company Report System Help**

**Available Commands:**
• **"Generate Report"** - Start the report generation process
• **"Export Data"** - Same as Generate Report
• **"Help"** - Show this help message

**Report Generation Process:**
1. I'll show you all available tables
2. You select which table to export
3. Choose specific columns or export all
4. Receive the CSV file via email

**Available Tables:**
• Users (MRs & Distributors)
• Orders
• Order Items  
• Products (Master)
• Customers
• Cart Items
• FOC Schemes
• Dealer Stock Details
• Pending Order Products
• Email Logs

Simply type **"generate report"** to get started!"""
            
            save_conversation(user.id, user_message, response)
            
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Generate Report', 'action': 'generate_report'}
                ]
            }), 200
        
        # Default response
        else:
            response = f"""I'm not sure how to help with that. As a company admin, I can help you:

• **Generate database reports** (type "generate report")
• **Export data to CSV** (type "export data")
• **Get help** (type "help")

What would you like to do?"""
            
            save_conversation(user.id, user_message, response)
            
            return jsonify({
                'response': response,
                'action_buttons': [
                    {'text': 'Generate Report', 'action': 'generate_report'},
                    {'text': 'Help', 'action': 'help'}
                ]
            }), 200
            
    except Exception as e:
        logger.error(f"Error handling company request: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'response': 'Sorry, I encountered an error. Please try again.',
            'action_buttons': [
                {'text': 'Generate Report', 'action': 'generate_report'}
            ]
        }), 500


@chatbot_bp.route('/company/select_table', methods=['POST'])
def select_company_table():
    """Handle company table selection for report generation"""
    try:
        from app.company_report_service import CompanyReportService
        
        data = request.json
        table_key = data.get('table_key')
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'company':
            return jsonify({'error': 'Unauthorized'}), 403
        
        report_service = CompanyReportService()
        available_tables = report_service.get_available_tables()
        
        if table_key not in available_tables:
            return jsonify({'error': 'Invalid table'}), 400
        
        table_info = available_tables[table_key]
        
        response = f"""**✅ Table Selected: {table_info['name']}**

This table has **{len(table_info['columns'])}** columns available.

**Column Selection:**

You can choose to:
• **Export all columns** (recommended for complete data)
• **Select specific columns** (for focused analysis)

**Available Columns:**

"""
        
        for col in table_info['columns']:
            response += f"• `{col}`\n"
        
        response += "\n**What would you like to do?**"
        
        save_conversation(user.id, f"Selected table: {table_key}", response)
        
        return jsonify({
            'response': response,
            'table_key': table_key,
            'table_name': table_info['name'],
            'columns': table_info['columns'],
            'show_column_selection': True,
            'action_buttons': []
        }), 200
        
    except Exception as e:
        logger.error(f"Error selecting table: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@chatbot_bp.route('/company/generate_report', methods=['POST'])
def generate_company_report():
    """Generate and send report to company email"""
    try:
        from app.company_report_service import CompanyReportService
        
        data = request.json
        table_key = data.get('table_key')
        selected_columns = data.get('selected_columns', [])
        user_id = session.get('user_id')
        
        if not user_id:
            return jsonify({'error': 'Not authenticated'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'company':
            return jsonify({'error': 'Unauthorized'}), 403
        
        report_service = CompanyReportService()
        
        # Generate report
        logger.info(f"Generating report for table: {table_key}, columns: {selected_columns}")
        report_result = report_service.generate_report(table_key, selected_columns)
        
        if not report_result.get('success'):
            return jsonify({
                'error': report_result.get('error', 'Report generation failed')
            }), 400
        
        # Send email
        email_result = report_service.send_report_email(
            user.email,
            table_key,
            report_result,
            selected_columns
        )
        
        if not email_result.get('success'):
            return jsonify({
                'error': email_result.get('error', 'Failed to send email')
            }), 500
        
        # Success response
        response = f"""**✅ Report Generated Successfully!**

Your report has been generated and sent to your email.

**Report Details:**
• **Table:** {report_result['table_name']}
• **Records:** {report_result['row_count']:,} rows
• **Columns:** {report_result['column_count']} columns
• **File:** {report_result['filename']}
• **Sent to:** {user.email}

📧 Please check your email inbox (and spam folder) for the report.

Would you like to generate another report?"""
        
        save_conversation(user.id, f"Generated report: {table_key}", response)
        
        return jsonify({
            'response': response,
            'success': True,
            'action_buttons': [
                {'text': 'Generate Another Report', 'action': 'generate_report'},
                {'text': 'Done', 'action': 'home'}
            ]
        }), 200
        
    except Exception as e:
        logger.error(f"Error generating report: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'error': str(e),
            'response': 'Sorry, there was an error generating your report. Please try again.'
        }), 500


# ============================================================================
# VOICE INTERACTION ENDPOINTS
# ============================================================================

@chatbot_bp.route('/api/voice/token', methods=['GET'])
def get_voice_token():
    """
    Get Azure Speech Service access token for client-side STT
    VOICE FEATURE IS DISABLED
    """
    # Voice feature disabled - always return error
    return jsonify({
        'error': 'Voice feature is currently disabled'
    }), 503
    
    # Original code commented out
    # try:
    #     speech_service = get_speech_service()
    #     
    #     if not speech_service.is_enabled():
    #         return jsonify({
    #             'error': 'Azure Speech Service is not configured'
    #         }), 503
    #     
    #     token = speech_service.get_access_token()
    #     
    #     if not token:
    #         return jsonify({
    #             'error': 'Failed to obtain access token'
    #         }), 500
    #     
    #     return jsonify({
    #         'token': token,
    #         'region': speech_service.speech_region
    #     }), 200
    #     
    # except Exception as e:
    #     logger.error(f"Error getting voice token: {str(e)}")
    #     return jsonify({
    #         'error': str(e)
    #     }), 500


@chatbot_bp.route('/api/voice/tts', methods=['POST'])
def text_to_speech():
    """
    Convert text to speech using Azure TTS
    VOICE FEATURE IS DISABLED
    """
    # Voice feature disabled - always return error
    return jsonify({
        'error': 'Voice feature is currently disabled'
    }), 503
    
    # Original code commented out - unreachable after return above
    # try:
    #     # Get JSON data
    #     try:
    #         data = request.get_json(force=True)
    #     except Exception as json_error:
    #         logger.error(f"Failed to parse JSON: {json_error}")
    #         return jsonify({
    #             'error': 'Invalid JSON data'
    #         }), 400
    #         
    #     if not data:
    #         return jsonify({
    #             'error': 'No JSON data provided'
    #         }), 400
    #         
    #     text = data.get('text', '')
    #     language = data.get('language', 'en')
    #     
    #     if not text or not text.strip():
    #         return jsonify({
    #             'error': 'No text provided'
    #         }), 400
    #     
    #     # Clean and limit text length (Azure TTS has limits)
    #     text = text.strip()
    #     if len(text) > 4000:
    #         text = text[:4000] + '...'
    #     
    #     # Get speech service
    #     try:
    #         speech_service = get_speech_service()
    #     except Exception as service_error:
    #         logger.error(f"Failed to get speech service: {service_error}")
    #         import traceback
    #         logger.error(traceback.format_exc())
    #         return jsonify({
    #             'error': f'Failed to initialize speech service: {str(service_error)}'
    #         }), 500
    #     
    #     if not speech_service or not speech_service.is_enabled():
    #         logger.warning("TTS requested but Azure Speech Service is not enabled")
    #         return jsonify({
    #             'error': 'Azure Speech Service is not configured'
    #         }), 503
    #     
    #     # Convert text to speech
    #     logger.info(f"Generating TTS for {len(text)} characters in language {language}")
    #     try:
    #         audio_data = speech_service.text_to_speech(text, language)
    #     except Exception as tts_error:
    #         logger.error(f"TTS generation failed: {tts_error}")
    #         import traceback
    #         logger.error(traceback.format_exc())
    #         return jsonify({
    #             'error': f'TTS generation failed: {str(tts_error)}'
    #         }), 500
    #     
    #     if not audio_data:
    #         logger.error("TTS generation returned None")
    #         return jsonify({
    #             'error': 'Failed to generate speech'
    #         }), 500
    #     
    #     # Ensure audio_data is at the beginning
    #     try:
    #         audio_data.seek(0)
    #     except Exception as seek_error:
    #         logger.error(f"Failed to seek audio data: {seek_error}")
    #         return jsonify({
    #             'error': f'Failed to process audio data: {str(seek_error)}'
    #         }), 500
    #     
    #     # Read all audio data into bytes (this is safe since we know it's not huge)
    #     try:
    #         audio_bytes = audio_data.read()
    #         audio_size = len(audio_bytes)
    #     except Exception as read_error:
    #         logger.error(f"Failed to read audio data: {read_error}")
    #         import traceback
    #         logger.error(traceback.format_exc())
    #         return jsonify({
    #             'error': f'Failed to read audio data: {str(read_error)}'
    #         }), 500
    #     
    #     if audio_size == 0:
    #         logger.error("TTS generated empty audio data")
    #         return jsonify({
    #             'error': 'Generated audio is empty'
    #         }), 500
    #     
    #     # Create response with audio data directly
    #     try:
    #         response = make_response(audio_bytes)
    #         response.headers['Content-Type'] = 'audio/mpeg'
    #         response.headers['Content-Length'] = str(audio_size)
    #         response.headers['Content-Disposition'] = 'inline; filename=speech.mp3'
    #         response.headers['Access-Control-Allow-Origin'] = '*'
    #         response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    #         response.headers['Pragma'] = 'no-cache'
    #         response.headers['Expires'] = '0'
    #         
    #         logger.info(f"TTS audio generated successfully, size: {audio_size} bytes")
    #         return response
    #     except Exception as response_error:
    #         logger.error(f"Failed to create response: {response_error}")
    #         import traceback
    #         logger.error(traceback.format_exc())
    #         return jsonify({
    #             'error': f'Failed to create response: {str(response_error)}'
    #         }), 500
    #     
    # except Exception as e:
    #     logger.error(f"Unexpected error in TTS endpoint: {str(e)}")
    #     import traceback
    #     error_trace = traceback.format_exc()
    #     logger.error(error_trace)
    #     
    #     # Return detailed error for debugging
    #     return jsonify({
    #         'error': str(e),
    #         'error_type': type(e).__name__,
    #         'details': error_trace.split('\n')[-10:] if len(error_trace) > 0 else []
    #     }), 500


@chatbot_bp.route('/api/voice/config', methods=['GET'])
def get_voice_config():
    """
    Get voice configuration (voices, languages, etc.)
    VOICE FEATURE IS DISABLED
    """
    # Voice feature disabled - always return disabled
    return jsonify({
        'enabled': False,
        'error': 'Voice feature is currently disabled'
    }), 200
    
    # Original code commented out
    # try:
    #     speech_service = get_speech_service()
    #     
    #     if not speech_service or not speech_service.is_enabled():
    #         return jsonify({
    #             'enabled': False,
    #             'error': 'Azure Speech Service is not configured'
    #         }), 200
    #     
    #     # Safely get attributes
    #     region = getattr(speech_service, 'speech_region', 'eastus')
    #     voice_map = getattr(speech_service, 'voice_map', {})
    #     
    #     return jsonify({
    #         'enabled': True,
    #         'region': region,
    #         'voices': voice_map,
    #         'languages': list(voice_map.keys()) if voice_map else []
    #     }), 200
    #     
    # except Exception as e:
    #     logger.error(f"Error getting voice config: {str(e)}")
    #     import traceback
    #     logger.error(traceback.format_exc())
    #     return jsonify({
    #         'enabled': False,
    #         'error': str(e)
    #     }), 200  # Return 200 instead of 500 to prevent frontend errors


@chatbot_bp.route('/api/delivery-partners', methods=['GET'])
def get_delivery_partners():
    """Get delivery partners for a given area (for dealer to select)"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'distributor':
            return jsonify({'error': 'Only distributors can view delivery partners'}), 403
        
        # Get delivery partners in the same area as the distributor
        area = user.area
        if not area:
            return jsonify({'error': 'Distributor area not set'}), 400
        
        logger.info(f"Fetching delivery partners for area: '{area}' (distributor: {user.name}, user_id: {user_id})")
        
        delivery_partners = User.query.filter_by(
            role='delivery_partner',
            area=area,
            is_active=True
        ).all()
        
        logger.info(f"Found {len(delivery_partners)} delivery partners in area '{area}'")
        
        partners_list = [{
            'id': dp.id,
            'unique_id': dp.unique_id,
            'name': dp.name,
            'email': dp.email,
            'phone': dp.phone,
            'area': dp.area
        } for dp in delivery_partners]
        
        if not partners_list:
            logger.warning(f"No delivery partners found for area '{area}'. Available areas: {[dp.area for dp in User.query.filter_by(role='delivery_partner', is_active=True).all()]}")
        
        return jsonify({
            'success': True,
            'delivery_partners': partners_list,
            'area': area  # Include area in response for debugging
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting delivery partners: {str(e)}")
        return jsonify({'error': 'Error getting delivery partners'}), 500


@chatbot_bp.route('/api/delivery-partner/orders', methods=['GET'])
def get_delivery_partner_orders():
    """Get all orders assigned to the logged-in delivery partner"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'delivery_partner':
            return jsonify({'error': 'Only delivery partners can view assigned orders'}), 403
        
        # Get orders assigned to this delivery partner that are not yet delivered
        # Include both confirmed and in_transit orders (but not delivered/cancelled)
        from sqlalchemy import or_, not_
        excluded_statuses = ['delivered', 'cancelled', 'completed']
        orders = Order.query.filter(
            or_(
                Order.delivery_partner_id == user_id,
                Order.delivery_partner_unique_id == user.unique_id
            )
        ).filter(
            ~Order.status.in_(excluded_statuses)
        ).order_by(Order.created_at.desc()).all()
        
        orders_list = []
        for order in orders:
            # Get customer details
            customer = None
            if order.customer_id:
                customer = Customer.query.get(order.customer_id)
            
            # Get order items
            order_items = OrderItem.query.filter_by(order_id=order.id).all()
            items = [{
                'product_code': item.product_code,
                'product_name': item.product_name,
                'quantity': item.adjusted_quantity if item.adjusted_quantity else item.quantity,
                'free_quantity': item.free_quantity or 0,
                'unit_price': item.unit_price
            } for item in order_items]
            
            orders_list.append({
                'order_id': order.order_id,
                'order_date': order.created_at.isoformat() if order.created_at else None,
                'total_amount': order.total_amount,
                'status': order.status or 'confirmed',
                'status_display': (order.status or 'confirmed').replace('_', ' ').title(),
                'customer': {
                    'name': customer.name if customer else 'N/A',
                    'phone': customer.phone if customer else 'N/A',
                    'address': customer.address if customer else 'Address not provided'
                },
                'items': items
            })
        
        return jsonify({
            'success': True,
            'orders': orders_list
        }), 200
        
    except Exception as e:
        logger.error(f"Error getting delivery partner orders: {str(e)}")
        return jsonify({'error': 'Error getting orders'}), 500


@chatbot_bp.route('/api/delivery-partner/mark-delivered', methods=['POST'])
def mark_order_delivered():
    """Mark an order as delivered by delivery partner"""
    try:
        user_id = session.get('user_id')
        if not user_id:
            return jsonify({'error': 'User not logged in'}), 401
        
        user = User.query.get(user_id)
        if not user or user.role != 'delivery_partner':
            return jsonify({'error': 'Only delivery partners can mark orders as delivered'}), 403
        
        data = request.get_json()
        if not data:
            return jsonify({'error': 'Invalid JSON or missing Content-Type header'}), 400
        
        order_id = data.get('order_id')
        if not order_id:
            return jsonify({'error': 'Order ID is required'}), 400
        
        # Validate order ID
        from app.input_validation import validate_and_sanitize_order_id
        order_id = validate_and_sanitize_order_id(order_id)
        if not order_id:
            return jsonify({'error': 'Invalid order ID format'}), 400
        
        # Mark order as delivered
        enhanced_order_service = get_enhanced_order_service()
        result = enhanced_order_service.mark_order_as_delivered(order_id, user_id)
        
        return jsonify(result), 200 if result['success'] else 400
        
    except Exception as e:
        logger.error(f"Error marking order as delivered: {str(e)}")
        return jsonify({'error': 'Error marking order as delivered'}), 500
